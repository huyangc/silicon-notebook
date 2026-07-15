"""Grid parsing for knowhow tables: xlsx/csv/markdown pipe tables -> ParsedGrid.

Pure functions only — no I/O beyond decoding the bytes handed in, no
repository/DB access. Callers (ingestion routes/services) own file I/O and
persistence; this module just turns bytes into a rectangular grid of strings
plus a best-effort guess of each column's semantic role.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path


class GridParseError(ValueError):
    """A grid could not be parsed into a valid table. Raised for: an empty
    header (empty file, or a first row with no column names); duplicate column
    names; zero data rows (nothing after the header, OR every data row blank);
    an unsupported file suffix; text (csv/md) that decodes as neither UTF-8
    (BOM-tolerant) nor GBK; or an xlsx openpyxl cannot open (corrupt / not a
    real .xlsx) or that contains no worksheet. Carries a user-facing Chinese
    message."""


@dataclass
class ParsedGrid:
    columns: list[str]
    rows: list[list[str]]


# Markdown separator-row cell, e.g. "---", ":--", "--:", ":-:" (GFM alignment markers).
_SEPARATOR_CELL_RE = re.compile(r":?-{3,}:?")
_PIPE_SPLIT_RE = re.compile(r"(?<!\\)\|")
_BR_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)


def parse_grid(filename: str, data: bytes) -> ParsedGrid:
    """Parse xlsx/xlsm/csv/md bytes into a ParsedGrid, dispatched by suffix.

    Raises GridParseError for an empty header, duplicate column names, zero
    data rows, an unsupported file suffix, text that decodes as neither
    UTF-8 (with optional BOM) nor GBK (csv/md), or xlsx bytes that are
    corrupted/not a valid Excel file. Data rows shorter than the header are
    padded with "" to header length; longer rows are truncated — neither
    case raises.
    """
    suffix = Path(filename).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        raw_rows = _extract_xlsx_rows(data)
    elif suffix == ".csv":
        raw_rows = _extract_csv_rows(data)
    elif suffix in (".md", ".markdown"):
        raw_rows = _extract_markdown_rows(data)
    else:
        raise GridParseError(f"不支持的表格文件类型：{suffix or filename}")
    return _build_grid(raw_rows)


def _build_grid(raw_rows: list[list[str]]) -> ParsedGrid:
    if not raw_rows or not raw_rows[0]:
        raise GridParseError("表格缺少表头：文件为空，或首行没有任何列名")

    columns = raw_rows[0]
    seen: set[str] = set()
    for name in columns:
        if name in seen:
            raise GridParseError(f"表头存在重复列名：{name!r}")
        seen.add(name)

    width = len(columns)
    normalized_rows: list[list[str]] = []
    for row in raw_rows[1:]:
        if len(row) < width:
            row = row + [""] * (width - len(row))
        elif len(row) > width:
            row = row[:width]
        # Drop rows with no content in ANY cell. A trailing formatted-but-empty
        # row is common in xlsx exports (openpyxl yields it as all-"" cells) and
        # would otherwise become a phantom data row — and, downstream, a phantom
        # case KO. Checked AFTER pad/truncate so a short all-blank row is caught
        # too; total_rows stays honest (only real rows counted).
        if not any(cell.strip() for cell in row):
            continue
        normalized_rows.append(row)

    if not normalized_rows:
        raise GridParseError("表格没有数据行：表头之后未找到任何数据")

    return ParsedGrid(columns=columns, rows=normalized_rows)


def _extract_xlsx_rows(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise GridParseError("无法读取 Excel 文件，请确认文件未损坏且为 .xlsx 格式") from exc
    try:
        # A workbook openpyxl opens but that carries no worksheet at all —
        # worksheets[0] would raise a bare IndexError; surface the same
        # friendly corrupt/invalid message instead.
        if not workbook.worksheets:
            raise GridParseError("无法读取 Excel 文件，请确认文件未损坏且为 .xlsx 格式")
        sheet = workbook.worksheets[0]
        return [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


def _decode_text(data: bytes) -> str:
    """Decode bytes as UTF-8 (BOM-tolerant), falling back to GBK.

    Chinese Excel exports CSV as GBK/ANSI by default rather than UTF-8, so
    the fallback is a deliberate accommodation for that common real-world
    case, not just a defensive catch-all.
    """
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    try:
        return data.decode("gbk")
    except UnicodeDecodeError as exc:
        raise GridParseError("无法识别文件编码，请将文件另存为 UTF-8 编码后重试") from exc


def _extract_csv_rows(data: bytes) -> list[list[str]]:
    text = _decode_text(data)
    return [list(row) for row in csv.reader(io.StringIO(text))]


def _extract_markdown_rows(data: bytes) -> list[list[str]]:
    text = _decode_text(data)
    rows: list[list[str]] = []
    for line in text.splitlines():
        stripped = line.strip()
        if len(stripped) < 2 or not (stripped.startswith("|") and stripped.endswith("|")):
            continue
        # Leading/trailing entries are artifacts of the outer pipe delimiters.
        cells = [_clean_md_cell(cell) for cell in _PIPE_SPLIT_RE.split(stripped)[1:-1]]
        if cells and all(_SEPARATOR_CELL_RE.fullmatch(cell) for cell in cells):
            continue  # alignment/separator row, not header or data
        rows.append(cells)
    return rows


def _clean_md_cell(raw: str) -> str:
    cell = raw.strip()
    cell = cell.replace("\\|", "|")
    cell = _BR_RE.sub("\n", cell)
    return cell


# Behavior-kind keyword table (knowhow-tables PR-2+3 Task 1, design doc §①
# "角色词表(2026-07-15 修订)"): the five domain-instance roles collapsed into
# domain-neutral kinds. Order = priority (procedure before entity, so a
# 「修复工具」-style double hit reads as a how-to column, mirroring the old
# table's priority-by-order convention).
KIND_KEYWORDS: list[tuple[str, tuple[str, ...]]] = [
    ("procedure", ("识别", "方法", "步骤", "分析", "修复")),
    ("entity", ("工具", "命令", "脚本")),
]

# Anchor (row-title column) NAME keywords — a separate vocabulary from the
# content kinds above, because "which column names the row" is a table-level
# identity question, not a content-type one. Matched casefolded; deliberately
# NO first-column fallback (design doc §① 旅行日志实证: a record-shaped table
# has no identity column, and guessing one manufactures phantom KG nodes).
ANCHOR_NAME_KEYWORDS: tuple[str, ...] = (
    "名称", "概念", "类型", "violation", "name", "type", "concept",
)


def guess_kinds(columns: list[str]) -> tuple[list[str], "int | None"]:
    """Best-effort header heuristics for the import/create wizards: each
    column's behavior kind (``procedure``/``entity``/``attribute`` — never
    ``anchor``, which is not a content kind) plus, SEPARATELY, the index of
    the first column whose NAME suggests it holds the row identity (the
    anchor/row-title suggestion), or ``None`` when nothing qualifies. Both
    are suggestions only — the user confirms/edits them in the wizard."""
    kinds: list[str] = []
    for name in columns:
        low = name.strip().casefold()
        for kind, keywords in KIND_KEYWORDS:
            if any(keyword in low for keyword in keywords):
                kinds.append(kind)
                break
        else:
            kinds.append("attribute")
    anchor_index: "int | None" = None
    for index, name in enumerate(columns):
        low = name.strip().casefold()
        if any(keyword in low for keyword in ANCHOR_NAME_KEYWORDS):
            anchor_index = index
            break
    return kinds, anchor_index


# transitional shim, removed by Task 3 (which rewires the preview endpoint to
# emit guessed_kind + anchor_suggestion directly): maps guess_kinds' output
# back onto the legacy wire strings the untouched PR-1 preview endpoint still
# returns, so the openapi golden stays byte-identical this task. The legacy
# identify/root_cause/fix distinction no longer exists — every procedure
# column maps to 'identify'.
_LEGACY_ROLE_BY_KIND = {
    "procedure": "identify",
    "entity": "tool",
    "attribute": "plain",
}


def guess_roles(columns: list[str]) -> list[str]:
    kinds, anchor_index = guess_kinds(columns)
    roles = [_LEGACY_ROLE_BY_KIND[kind] for kind in kinds]
    if anchor_index is not None:
        roles[anchor_index] = "concept"
    return roles
