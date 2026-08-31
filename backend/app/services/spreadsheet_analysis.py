"""Deterministic spreadsheet compilation and bounded reasoning-time analysis."""
from __future__ import annotations

import json
import math
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from statistics import fmean
from time import perf_counter
from typing import Any, Iterable, Sequence

from app.models.ask import (
    Citation,
    SpreadsheetAnalysisResult,
    SpreadsheetResultRow,
    StructuredResultColumn,
    StructuredResultCoverage,
    TraceStep,
)
from app.repositories.analysis_artifacts import AnalysisArtifactStore
from app.services.cancellation import AskCancelled, raise_if_cancelled


SPREADSHEET_SUFFIXES = frozenset({".xlsx", ".xlsm", ".xls"})
PLAN_OPERATIONS = frozenset({"profile", "aggregate", "top", "filter"})
AGGREGATIONS = frozenset({"sum", "avg", "min", "max", "count"})
FILTER_OPERATORS = frozenset({"eq", "ne", "gt", "gte", "lt", "lte", "contains"})
_ANALYTICAL_TERMS = (
    "excel", "工作簿", "表格", "sheet", "worksheet", "数据", "分析", "统计",
    "汇总", "合计", "总和", "平均", "最大", "最小", "中位", "趋势", "排名",
    "top", "筛选", "过滤", "分组", "透视", "占比", "同比", "环比", "异常",
    "缺失", "重复", "列", "行", "sum", "average", "count", "group", "filter",
)


class SpreadsheetCompileError(RuntimeError):
    def __init__(self, code: str, summary: str) -> None:
        super().__init__(code)
        self.code = code
        self.summary = summary


def _json_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, str)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _display_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return format(value, ".12g") if math.isfinite(value) else str(value)
    return str(value)


def _json_bytes(value: Any) -> int:
    return len(
        json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    )


def _numeric(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    if isinstance(value, str):
        candidate = value.strip().replace(",", "")
        if candidate.endswith("%"):
            try:
                return float(candidate[:-1]) / 100
            except ValueError:
                return None
        try:
            parsed = float(candidate)
        except ValueError:
            return None
        return parsed if math.isfinite(parsed) else None
    return None


def _column_name(index: int) -> str:
    value = index + 1
    text = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        text = chr(65 + remainder) + text
    return text


def _unique_headers(values: Sequence[Any]) -> list[str]:
    seen: Counter[str] = Counter()
    headers: list[str] = []
    for index, value in enumerate(values):
        base = _display_cell(value).strip() or f"列 {_column_name(index)}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    return headers


def _header_index(rows: Sequence[dict[str, Any]]) -> int:
    candidates: list[tuple[int, int, int]] = []
    for index, row in enumerate(rows[:20]):
        values = row["cells"]
        populated = sum(value not in (None, "") for value in values)
        if populated < 2:
            continue
        textual = sum(
            isinstance(value, str) and bool(value.strip()) for value in values
        )
        has_following = any(
            any(value not in (None, "") for value in following["cells"])
            for following in rows[index + 1:index + 4]
        )
        if has_following:
            candidates.append((textual * 4 + populated, -index, index))
    if not candidates:
        raise SpreadsheetCompileError(
            "SPREADSHEET_HEADER_NOT_FOUND",
            "工作表中没有识别到可分析的表头和数据区域。",
        )
    return max(candidates)[2]


def _looks_like_header(row: dict[str, Any]) -> bool:
    populated = [value for value in row["cells"] if value not in (None, "")]
    return (
        len(populated) >= 2
        and all(isinstance(value, str) and bool(value.strip()) for value in populated)
    )


class SpreadsheetAnalysisService:
    """Compile once during ingestion; execute a whitelist plan during Ask."""

    def __init__(
        self,
        *,
        artifacts: AnalysisArtifactStore,
        settings: Any,
        event_log: Any,
        now: Any,
    ) -> None:
        self.artifacts = artifacts
        self.settings = settings
        self.event_log = event_log
        self.now = now

    @staticmethod
    def supports(file_name: str) -> bool:
        return Path(file_name or "").suffix.lower() in SPREADSHEET_SUFFIXES

    def compile_source(
        self,
        source: Any,
        *,
        notebook_name: str,
        owner_id: str,
        row_element_ids: dict[tuple[str, int], str],
    ) -> bool:
        """Build or replace a source snapshot without changing source status."""
        if not self.settings.spreadsheet_analysis_enabled:
            return False
        if not self.supports(getattr(source, "file_name", "")):
            return False
        started = perf_counter()
        try:
            manifest = self._compile_workbook(source, row_element_ids=row_element_ids)
            self.artifacts.save_spreadsheet_manifest(
                source.notebook_id, source.id, manifest
            )
            self.artifacts.resolve_issue(
                source.notebook_id,
                source.id,
                "spreadsheet_analysis",
                resolved_at=self.now(),
            )
            self.event_log.emit({
                "kind": "spreadsheet_analysis",
                "source_id": source.id,
                "notebook_id": source.notebook_id,
                "stage": "compile",
                "status": "ready",
                "latency_ms": round((perf_counter() - started) * 1000),
                "sheets": len(manifest["sheets"]),
                "cells": manifest["cell_count"],
            })
            return True
        except SpreadsheetCompileError as exc:
            self._record_compile_issue(
                source, notebook_name=notebook_name, owner_id=owner_id,
                code=exc.code, summary=exc.summary,
            )
        except Exception as exc:  # noqa: BLE001 - optional lane, source parsing survives
            self.event_log.logger.warning(
                "spreadsheet analysis compilation failed (%s)", type(exc).__name__
            )
            self._record_compile_issue(
                source, notebook_name=notebook_name, owner_id=owner_id,
                code=f"SPREADSHEET_COMPILE_{type(exc).__name__.upper()}",
                summary="工作簿结构超出当前专业分析能力，普通文本检索仍可继续使用。",
            )
        self.artifacts.delete_spreadsheet_manifest(source.notebook_id, source.id)
        return False

    def _record_compile_issue(
        self,
        source: Any,
        *,
        notebook_name: str,
        owner_id: str,
        code: str,
        summary: str,
    ) -> None:
        self.artifacts.record_issue(
            notebook_id=source.notebook_id,
            notebook_name=notebook_name,
            owner_id=owner_id,
            source_id=source.id,
            source_title=source.title,
            file_name=source.file_name,
            source_type=source.type,
            category="spreadsheet_analysis",
            code=code,
            summary=summary,
            occurred_at=self.now(),
            source_path=source.file_path,
            source_hash=getattr(source, "file_hash", "") or "",
            archive_file=True,
        )
        self.event_log.emit({
            "kind": "analysis_issue",
            "source_id": source.id,
            "notebook_id": source.notebook_id,
            "category": "spreadsheet_analysis",
            "code": code,
            "status": "open",
        })

    def _compile_workbook(
        self, source: Any, *, row_element_ids: dict[tuple[str, int], str]
    ) -> dict[str, Any]:
        path = Path(source.file_path)
        if not path.is_file():
            raise SpreadsheetCompileError(
                "SPREADSHEET_FILE_MISSING", "工作簿原文件不存在，无法建立分析快照。"
            )
        if path.stat().st_size > self.settings.source_upload_max_bytes:
            raise SpreadsheetCompileError(
                "SPREADSHEET_FILE_TOO_LARGE", "工作簿超过当前部署允许的单文件分析上限。"
            )
        suffix = Path(source.file_name or path.name).suffix.lower()
        if suffix == ".xls" and not zipfile.is_zipfile(path):
            sheets = self._compile_xls(path, row_element_ids)
        else:
            sheets = self._compile_ooxml(path, row_element_ids)
        if not sheets:
            raise SpreadsheetCompileError(
                "SPREADSHEET_EMPTY", "工作簿中没有可分析的数据区域。"
            )
        cell_count = sum(
            len(sheet["rows"]) * len(sheet["headers"]) for sheet in sheets
        )
        return {
            "version": 1,
            "source_id": source.id,
            "notebook_id": source.notebook_id,
            "source_title": source.title,
            "source_file_name": source.file_name,
            "source_hash": getattr(source, "file_hash", "") or "",
            "compiled_at": self.now(),
            "cell_count": cell_count,
            "sheets": sheets,
        }

    def _compile_cell(self, value: Any) -> Any:
        normalized = _json_cell(value)
        if (
            isinstance(normalized, str)
            and len(normalized) > self.settings.spreadsheet_analysis_max_cell_chars
        ):
            raise SpreadsheetCompileError(
                "SPREADSHEET_CELL_TOO_LONG",
                "工作簿包含超长单元格，当前无法在不截断内容的情况下可靠分析。",
            )
        return normalized

    def _compile_ooxml(
        self, path: Path, row_element_ids: dict[tuple[str, int], str]
    ) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise SpreadsheetCompileError(
                "SPREADSHEET_DEPENDENCY_MISSING", "服务器未安装工作簿分析依赖。"
            ) from exc
        formula_handle = None
        value_handle = None
        formula_book = None
        value_book = None
        try:
            # Read-only workbooks stream lazily from their file handles, so the
            # handles stay open through the full worksheet walk below.
            formula_handle = path.open("rb")
            value_handle = path.open("rb")
            formula_book = load_workbook(
                formula_handle, read_only=True, data_only=False, keep_links=False
            )
            value_book = load_workbook(
                value_handle, read_only=True, data_only=True, keep_links=False
            )
        except Exception as exc:
            if formula_book is not None:
                formula_book.close()
            if value_book is not None:
                value_book.close()
            if formula_handle is not None:
                formula_handle.close()
            if value_handle is not None:
                value_handle.close()
            raise SpreadsheetCompileError(
                "SPREADSHEET_INVALID_OOXML", "无法读取工作簿；文件可能损坏、加密或格式不兼容。"
            ) from exc
        try:
            if len(formula_book.worksheets) > self.settings.spreadsheet_analysis_max_sheets:
                raise SpreadsheetCompileError(
                    "SPREADSHEET_TOO_MANY_SHEETS", "工作簿的工作表数量超过当前分析上限。"
                )
            dimension_cells = sum(
                int(sheet.max_row or 0) * int(sheet.max_column or 0)
                for sheet in formula_book.worksheets
            )
            if dimension_cells > self.settings.spreadsheet_analysis_max_cells:
                raise SpreadsheetCompileError(
                    "SPREADSHEET_TOO_MANY_CELLS", "工作簿数据区域超过当前分析单元格上限。"
                )
            sheets: list[dict[str, Any]] = []
            value_by_name = {sheet.title: sheet for sheet in value_book.worksheets}
            for formula_sheet in formula_book.worksheets:
                value_sheet = value_by_name.get(formula_sheet.title)
                if value_sheet is None:
                    continue
                rows: list[dict[str, Any]] = []
                formula_count = 0
                unresolved_count = 0
                value_rows = value_sheet.iter_rows(values_only=True)
                for row_number, formula_row in enumerate(
                    formula_sheet.iter_rows(values_only=False), start=1
                ):
                    try:
                        cached_values = next(value_rows)
                    except StopIteration:
                        cached_values = ()
                    width = max(len(formula_row), len(cached_values))
                    values: list[Any] = []
                    for column in range(width):
                        formula_cell = formula_row[column] if column < len(formula_row) else None
                        formula_value = getattr(formula_cell, "value", None)
                        cached = cached_values[column] if column < len(cached_values) else None
                        if getattr(formula_cell, "data_type", "") == "f":
                            formula_count += 1
                            if cached is None:
                                unresolved_count += 1
                            values.append(self._compile_cell(cached))
                        else:
                            values.append(self._compile_cell(formula_value))
                    while values and values[-1] is None:
                        values.pop()
                    if not any(value not in (None, "") for value in values):
                        continue
                    rows.append({
                        "row_number": row_number,
                        "element_id": row_element_ids.get(
                            (formula_sheet.title, row_number), ""
                        ),
                        "cells": values,
                    })
                if rows:
                    sheets.append(self._sheet_manifest(
                        formula_sheet.title,
                        "hidden" if formula_sheet.sheet_state != "visible" else "visible",
                        rows,
                        formula_count=formula_count,
                        unresolved_count=unresolved_count,
                    ))
            return sheets
        finally:
            if formula_book is not None:
                formula_book.close()
            if value_book is not None:
                value_book.close()
            if formula_handle is not None:
                formula_handle.close()
            if value_handle is not None:
                value_handle.close()

    def _compile_xls(
        self, path: Path, row_element_ids: dict[tuple[str, int], str]
    ) -> list[dict[str, Any]]:
        try:
            import xlrd
        except ImportError as exc:
            raise SpreadsheetCompileError(
                "SPREADSHEET_DEPENDENCY_MISSING", "服务器未安装旧版 Excel 分析依赖。"
            ) from exc
        try:
            book = xlrd.open_workbook(str(path), on_demand=True)
        except Exception as exc:
            raise SpreadsheetCompileError(
                "SPREADSHEET_INVALID_XLS", "无法读取旧版 Excel 文件；文件可能损坏或加密。"
            ) from exc
        try:
            if book.nsheets > self.settings.spreadsheet_analysis_max_sheets:
                raise SpreadsheetCompileError(
                    "SPREADSHEET_TOO_MANY_SHEETS", "工作簿的工作表数量超过当前分析上限。"
                )
            dimension_cells = sum(
                book.sheet_by_index(index).nrows * book.sheet_by_index(index).ncols
                for index in range(book.nsheets)
            )
            if dimension_cells > self.settings.spreadsheet_analysis_max_cells:
                raise SpreadsheetCompileError(
                    "SPREADSHEET_TOO_MANY_CELLS", "工作簿数据区域超过当前分析单元格上限。"
                )
            sheets: list[dict[str, Any]] = []
            for index in range(book.nsheets):
                sheet = book.sheet_by_index(index)
                rows: list[dict[str, Any]] = []
                for row_index in range(sheet.nrows):
                    values = [
                        self._compile_cell(value) for value in sheet.row_values(row_index)
                    ]
                    while values and values[-1] in (None, ""):
                        values.pop()
                    if not any(value not in (None, "") for value in values):
                        continue
                    rows.append({
                        "row_number": row_index + 1,
                        "element_id": row_element_ids.get((sheet.name, row_index + 1), ""),
                        "cells": values,
                    })
                if rows:
                    sheets.append(self._sheet_manifest(
                        sheet.name, "visible", rows,
                        formula_count=0, unresolved_count=0,
                    ))
                book.unload_sheet(index)
            return sheets
        finally:
            book.release_resources()

    def _sheet_manifest(
        self,
        name: str,
        state: str,
        rows: list[dict[str, Any]],
        *,
        formula_count: int,
        unresolved_count: int,
    ) -> dict[str, Any]:
        header_index = _header_index(rows)
        for index in range(header_index + 1, len(rows) - 1):
            previous = rows[index - 1]
            current = rows[index]
            if (
                current["row_number"] > previous["row_number"] + 1
                and _looks_like_header(current)
            ):
                raise SpreadsheetCompileError(
                    "SPREADSHEET_MULTIPLE_REGIONS",
                    "工作表中识别到多个由空白行分隔的数据区域，当前无法可靠选择分析范围。",
                )
        width = max(len(row["cells"]) for row in rows[header_index:])
        headers = _unique_headers(
            [
                rows[header_index]["cells"][column]
                if column < len(rows[header_index]["cells"]) else None
                for column in range(width)
            ]
        )
        data_rows = rows[header_index + 1:]
        warnings: list[str] = []
        if header_index:
            warnings.append("表头上方存在标题或说明行，分析从识别出的表头开始。")
        if state != "visible":
            warnings.append("该工作表处于隐藏状态。")
        if unresolved_count:
            warnings.append("部分公式没有缓存结果；相关单元格按空值处理。")
        first_row = rows[header_index]["row_number"]
        last_row = rows[-1]["row_number"]
        return {
            "name": name,
            "state": state,
            "header_row": first_row,
            "headers": headers,
            "range": f"A{first_row}:{_column_name(width - 1)}{last_row}",
            "rows": data_rows,
            "formula_cells": formula_count,
            "unresolved_formula_cells": unresolved_count,
            "warnings": warnings,
        }

    def analyze(
        self,
        *,
        notebook_id: str,
        source_ids: Sequence[str],
        question: str,
        planner_client: Any,
        cancel_event: Any = None,
    ) -> tuple[list[SpreadsheetAnalysisResult], TraceStep | None]:
        if not self.settings.spreadsheet_analysis_enabled:
            return [], None
        if not any(term in question.lower() for term in _ANALYTICAL_TERMS):
            return [], None
        manifests = [
            manifest
            for source_id in source_ids
            if (manifest := self.artifacts.load_spreadsheet_manifest(
                notebook_id, source_id
            )) is not None
        ]
        if not manifests:
            return [], None
        raise_if_cancelled(cancel_event)
        started = perf_counter()
        plan = self._plan(question, manifests, planner_client, cancel_event)
        if plan is None:
            return [], None
        raise_if_cancelled(cancel_event)
        result = self._execute(plan, manifests)
        if result is None:
            return [], None
        duration_ms = round((perf_counter() - started) * 1000)
        trace = TraceStep(
            step_type="spreadsheet",
            summary=f"已分析 {result.source_title} · {result.sheet}",
            detail={
                "source_id": result.source_id,
                "sheet": result.sheet,
                "range": result.range,
                "operation": result.operation,
                "scanned_rows": result.coverage.scanned_rows,
                "returned_rows": result.coverage.returned_rows,
                "complete": result.coverage.complete,
            },
            duration_ms=duration_ms,
        )
        return [result], trace

    def _plan(
        self,
        question: str,
        manifests: Sequence[dict[str, Any]],
        client: Any,
        cancel_event: Any,
    ) -> dict[str, Any] | None:
        if getattr(client, "configured", False):
            catalog, planning_manifests, catalog_complete = (
                self._bounded_planner_catalog(manifests)
            )
            if not catalog or self._default_plan(planning_manifests) is None:
                fallback = self._default_plan(manifests)
                if fallback is not None:
                    fallback["__planner_catalog_complete"] = False
                return fallback
            prompt = (
                "为电子表格问题选择一个受限的确定性执行计划。只能使用给定 source_id、"
                "sheet 和列名；不得生成公式、代码或 SQL。operation 只能是 profile、"
                "aggregate、top、filter。aggregate 使用 aggregation=sum|avg|min|max|count，"
                "可选 measure/group_by；top 使用 sort_by/order=asc|desc；filter 可选 filters，"
                "每项 operator=eq|ne|gt|gte|lt|lte|contains。无法可靠匹配时返回 operation=profile。\n"
                f"问题：{question}\n可用工作簿：{json.dumps(catalog, ensure_ascii=False)}"
            )
            try:
                raw = client.chat_json(
                    [{"role": "user", "content": prompt}],
                    '{"source_id":"id","sheet":"name","operation":"profile",'
                    '"aggregation":"count","measure":"","group_by":"",'
                    '"sort_by":"","order":"desc","filters":[],"columns":[]}',
                    timeout=self.settings.spreadsheet_analysis_planner_timeout_seconds,
                    max_tokens=500,
                    cancel_event=cancel_event,
                )
                candidate = json.loads(raw)
                if isinstance(candidate, dict):
                    normalized = self._validate_plan(candidate, planning_manifests)
                    normalized["__planner_catalog_complete"] = catalog_complete
                    return normalized
            except AskCancelled:
                raise
            except Exception as exc:  # noqa: BLE001 - optional planner is fail-open
                self.event_log.logger.warning(
                    "spreadsheet planner failed (%s)", type(exc).__name__
                )
            fallback = self._default_plan(planning_manifests)
            if fallback is not None:
                fallback["__planner_catalog_complete"] = catalog_complete
            return fallback
        return self._default_plan(manifests)

    def _bounded_planner_catalog(
        self, manifests: Sequence[dict[str, Any]]
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
        """Project whole catalog fields without ever slicing user-authored text."""
        limit = self.settings.spreadsheet_analysis_planner_catalog_bytes
        catalog: list[dict[str, Any]] = []
        projected: list[dict[str, Any]] = []
        complete = True
        stop = False
        used_bytes = _json_bytes(catalog)
        for manifest in manifests:
            catalog_manifest = {
                "source_id": manifest["source_id"],
                "title": manifest.get("source_title", ""),
                "sheets": [],
            }
            projected_manifest = {
                "source_id": manifest["source_id"],
                "sheets": [],
            }
            if _json_bytes([*catalog, catalog_manifest]) > limit:
                complete = False
                break
            catalog.append(catalog_manifest)
            projected.append(projected_manifest)
            used_bytes = _json_bytes(catalog)
            for sheet in manifest.get("sheets", []):
                catalog_sheet = {
                    "name": sheet["name"],
                    "headers": [],
                    "rows": len(sheet["rows"]),
                }
                projected_sheet = {"name": sheet["name"], "headers": []}
                catalog_manifest["sheets"].append(catalog_sheet)
                projected_manifest["sheets"].append(projected_sheet)
                if _json_bytes(catalog) > limit:
                    catalog_manifest["sheets"].pop()
                    projected_manifest["sheets"].pop()
                    complete = False
                    stop = True
                    break
                used_bytes = _json_bytes(catalog)
                for header in sheet.get("headers", []):
                    header_bytes = _json_bytes(header)
                    added_bytes = header_bytes + (1 if catalog_sheet["headers"] else 0)
                    if used_bytes + added_bytes > limit:
                        complete = False
                        stop = True
                        break
                    catalog_sheet["headers"].append(header)
                    projected_sheet["headers"].append(header)
                    used_bytes += added_bytes
                if stop:
                    break
            if stop:
                break
        return catalog, projected, complete

    def _default_plan(self, manifests: Sequence[dict[str, Any]]) -> dict[str, Any] | None:
        first = next(
            (
                (manifest, sheet)
                for manifest in manifests
                for sheet in manifest.get("sheets", [])
            ),
            None,
        )
        if first is None:
            return None
        manifest, sheet = first
        return {
            "source_id": manifest["source_id"],
            "sheet": sheet["name"],
            "operation": "profile",
            "aggregation": "count",
            "measure": "",
            "group_by": "",
            "sort_by": "",
            "order": "desc",
            "filters": [],
            "columns": [],
        }

    def _validate_plan(
        self, plan: dict[str, Any], manifests: Sequence[dict[str, Any]]
    ) -> dict[str, Any]:
        fallback = self._default_plan(manifests)
        if fallback is None:
            return {}
        source_id = str(plan.get("source_id") or "")
        manifest = next((row for row in manifests if row["source_id"] == source_id), None)
        if manifest is None:
            manifest = next(iter(manifests))
            source_id = manifest["source_id"]
        sheet_name = str(plan.get("sheet") or "")
        sheet = next(
            (row for row in manifest.get("sheets", []) if row["name"] == sheet_name),
            None,
        )
        if sheet is None:
            sheet = next(iter(manifest.get("sheets", [])), None)
        if sheet is None:
            return fallback
        operation = str(plan.get("operation") or "profile").lower()
        if operation not in PLAN_OPERATIONS:
            operation = "profile"
        aggregation = str(plan.get("aggregation") or "count").lower()
        if aggregation not in AGGREGATIONS:
            aggregation = "count"
        headers = sheet["headers"]
        normalized = {
            "source_id": source_id,
            "sheet": sheet["name"],
            "operation": operation,
            "aggregation": aggregation,
            "measure": self._resolve_header(str(plan.get("measure") or ""), headers),
            "group_by": self._resolve_header(str(plan.get("group_by") or ""), headers),
            "sort_by": self._resolve_header(str(plan.get("sort_by") or ""), headers),
            "order": "asc" if str(plan.get("order") or "").lower() == "asc" else "desc",
            "columns": [
                resolved
                for value in plan.get("columns", []) if isinstance(value, str)
                if (resolved := self._resolve_header(value, headers))
            ][:20],
            "filters": [],
        }
        for item in plan.get("filters", [])[:8] if isinstance(plan.get("filters"), list) else []:
            if not isinstance(item, dict):
                continue
            column = self._resolve_header(str(item.get("column") or ""), headers)
            operator = str(item.get("operator") or "eq").lower()
            if column and operator in FILTER_OPERATORS:
                normalized["filters"].append({
                    "column": column,
                    "operator": operator,
                    "value": _json_cell(item.get("value")),
                })
        if operation == "aggregate" and aggregation != "count" and not normalized["measure"]:
            normalized["operation"] = "profile"
        if operation == "top" and not normalized["sort_by"]:
            normalized["operation"] = "profile"
        return normalized

    @staticmethod
    def _resolve_header(value: str, headers: Sequence[str]) -> str:
        wanted = value.strip().casefold()
        if not wanted:
            return ""
        exact = next((header for header in headers if header.casefold() == wanted), "")
        if exact:
            return exact
        matches = [header for header in headers if wanted in header.casefold()]
        return matches[0] if len(matches) == 1 else ""

    def _execute(
        self, plan: dict[str, Any], manifests: Sequence[dict[str, Any]]
    ) -> SpreadsheetAnalysisResult | None:
        manifest = next(
            (row for row in manifests if row["source_id"] == plan.get("source_id")), None
        )
        if manifest is None:
            return None
        sheet = next(
            (row for row in manifest.get("sheets", []) if row["name"] == plan.get("sheet")),
            None,
        )
        if sheet is None:
            return None
        headers = sheet["headers"]
        input_rows = [self._row_dict(row, headers) for row in sheet["rows"]]
        filtered = [row for row in input_rows if self._matches(row, plan.get("filters", []))]
        operation = plan["operation"]
        if operation == "aggregate":
            output = self._aggregate(filtered, plan)
        elif operation == "top":
            output = self._top(filtered, plan)
        elif operation == "filter":
            output = self._tabular(filtered, plan.get("columns", []))
        else:
            output = self._profile(filtered, headers)
        output_rows, output_headers = output
        citation = self._citation(manifest, sheet, filtered or input_rows)
        delivered, delivered_headers, complete, truncated_reason, budget_warnings = (
            self._bounded_result_preview(
                output_rows, output_headers, citation=citation
            )
        )
        warnings = list(sheet.get("warnings", []))
        warnings.extend(budget_warnings)
        if not plan.get("__planner_catalog_complete", True):
            warnings.append(
                "用于模型选表和选列的规划目录已按内容量上限做有界投影；"
                "完整工作表仍由本地确定性执行器扫描。"
            )
        rows = [
            SpreadsheetResultRow(
                position=index,
                cells=row,
                citation=citation if index == 1 else None,
            )
            for index, row in enumerate(delivered, start=1)
        ]
        return SpreadsheetAnalysisResult(
            source_id=manifest["source_id"],
            source_title=manifest.get("source_title", ""),
            source_file_name=manifest.get("source_file_name", ""),
            sheet=sheet["name"],
            range=sheet["range"],
            operation=operation,
            columns=[
                StructuredResultColumn(id=header, name=header)
                for header in delivered_headers
            ],
            rows=rows,
            coverage=StructuredResultCoverage(
                total_rows=len(output_rows),
                scanned_rows=len(input_rows),
                returned_rows=len(delivered),
                complete=complete,
                truncated_reason="" if complete else truncated_reason,
                overflow_semantics="explicit_partial" if not complete else "",
            ),
            formula_cells=int(sheet.get("formula_cells") or 0),
            unresolved_formula_cells=int(sheet.get("unresolved_formula_cells") or 0),
            warnings=warnings,
        )

    def _bounded_result_preview(
        self,
        output_rows: Sequence[dict[str, Any]],
        output_headers: Sequence[str],
        *,
        citation: Citation | None,
    ) -> tuple[list[dict[str, str]], list[str], bool, str, list[str]]:
        """Keep whole values while bounding rows, cells, and serialized table bytes."""
        row_cap = self.settings.spreadsheet_analysis_result_rows
        cell_cap = self.settings.spreadsheet_analysis_result_cells
        byte_cap = self.settings.spreadsheet_analysis_result_bytes
        citation_payload = (
            citation.model_dump(mode="json", exclude_none=True) if citation else None
        )

        def payload(headers: Sequence[str], rows: Sequence[dict[str, str]]) -> dict[str, Any]:
            wire_rows = []
            for index, row in enumerate(rows, start=1):
                item: dict[str, Any] = {"position": index, "cells": row}
                if index == 1 and citation_payload is not None:
                    item["citation"] = citation_payload
                wire_rows.append(item)
            return {
                "columns": [{"id": header, "name": header} for header in headers],
                "rows": wire_rows,
            }

        def cell_bytes(header: str, value: str, *, has_previous: bool) -> int:
            return (
                _json_bytes(header)
                + 1
                + _json_bytes(value)
                + (1 if has_previous else 0)
            )

        selected_headers: list[str] = []
        byte_limited = False
        cell_limited = False
        first_row = output_rows[0] if output_rows else None
        reserved_rows: list[dict[str, str]] = [{}] if first_row is not None else []
        reserved_bytes = _json_bytes(payload([], reserved_rows))
        for header in output_headers:
            displayed = _display_cell(first_row.get(header)) if first_row is not None else ""
            column_bytes = _json_bytes({"id": header, "name": header}) + (
                1 if selected_headers else 0
            )
            value_bytes = (
                cell_bytes(header, displayed, has_previous=bool(selected_headers))
                if first_row is not None else 0
            )
            candidate_bytes = reserved_bytes + column_bytes + value_bytes
            reserved_cells = (len(selected_headers) + 1) * (
                2 if first_row is not None else 1
            )
            if reserved_cells > cell_cap:
                cell_limited = True
                break
            if candidate_bytes > byte_cap:
                byte_limited = True
                break
            selected_headers.append(header)
            if first_row is not None:
                reserved_rows[0][header] = displayed
            reserved_bytes = candidate_bytes

        if not selected_headers and output_headers:
            first_header = output_headers[0]
            if _json_bytes(payload([first_header], [])) <= byte_cap:
                selected_headers.append(first_header)
            else:
                byte_limited = True

        delivered: list[dict[str, str]] = []
        delivered_bytes = _json_bytes(payload(selected_headers, delivered))
        candidate_rows = output_rows[:row_cap] if selected_headers else []
        for row in candidate_rows:
            rendered = {
                header: _display_cell(row.get(header)) for header in selected_headers
            }
            if len(selected_headers) + (len(delivered) + 1) * len(selected_headers) > cell_cap:
                cell_limited = True
                break
            row_payload: dict[str, Any] = {
                "position": len(delivered) + 1,
                "cells": rendered,
            }
            if not delivered and citation_payload is not None:
                row_payload["citation"] = citation_payload
            added_bytes = _json_bytes(row_payload) + (1 if delivered else 0)
            if delivered_bytes + added_bytes > byte_cap:
                byte_limited = True
                break
            delivered.append(rendered)
            delivered_bytes += added_bytes

        columns_complete = len(selected_headers) == len(output_headers)
        rows_complete = len(delivered) == len(output_rows)
        complete = columns_complete and rows_complete
        warnings: list[str] = []
        if not columns_complete:
            warnings.append(
                f"计算结果共 {len(output_headers)} 列，受单元格或内容量上限约束，"
                f"仅返回前 {len(selected_headers)} 列。"
            )
        if not rows_complete:
            warnings.append(
                f"计算结果共 {len(output_rows)} 行，受行数、单元格或内容量上限约束，"
                f"仅返回前 {len(delivered)} 行。"
            )
        if byte_limited:
            reason = "payload_limit"
        elif cell_limited or not columns_complete:
            reason = "budget"
        else:
            reason = "row_limit"
        return delivered, selected_headers, complete, reason, warnings

    @staticmethod
    def _row_dict(row: dict[str, Any], headers: Sequence[str]) -> dict[str, Any]:
        cells = row.get("cells", [])
        result = {
            header: cells[index] if index < len(cells) else None
            for index, header in enumerate(headers)
        }
        result["__element_id"] = row.get("element_id", "")
        result["__row_number"] = row.get("row_number", 0)
        return result

    def _matches(self, row: dict[str, Any], filters: Iterable[dict[str, Any]]) -> bool:
        for item in filters:
            left = row.get(item["column"])
            right = item.get("value")
            operator = item["operator"]
            left_number, right_number = _numeric(left), _numeric(right)
            if operator == "contains":
                matched = str(right or "").casefold() in str(left or "").casefold()
            elif operator in {"gt", "gte", "lt", "lte"}:
                if left_number is None or right_number is None:
                    return False
                matched = {
                    "gt": left_number > right_number,
                    "gte": left_number >= right_number,
                    "lt": left_number < right_number,
                    "lte": left_number <= right_number,
                }[operator]
            else:
                matched = str(left or "").casefold() == str(right or "").casefold()
                if operator == "ne":
                    matched = not matched
            if not matched:
                return False
        return True

    @staticmethod
    def _profile(
        rows: Sequence[dict[str, Any]], headers: Sequence[str]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        output: list[dict[str, Any]] = []
        for header in headers:
            values = [row.get(header) for row in rows]
            nonempty = [value for value in values if value not in (None, "")]
            numeric = [number for value in nonempty if (number := _numeric(value)) is not None]
            inferred = "数值" if nonempty and len(numeric) / len(nonempty) >= 0.8 else "文本"
            output.append({
                "列": header,
                "类型": inferred,
                "非空": len(nonempty),
                "缺失": len(values) - len(nonempty),
                "去重值": len({str(value) for value in nonempty}),
                "最小值": min(numeric) if numeric else "",
                "最大值": max(numeric) if numeric else "",
                "平均值": fmean(numeric) if numeric else "",
            })
        columns = ["列", "类型", "非空", "缺失", "去重值", "最小值", "最大值", "平均值"]
        return output, columns

    @staticmethod
    def _aggregate(
        rows: Sequence[dict[str, Any]], plan: dict[str, Any]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        group_by = plan.get("group_by") or ""
        aggregation = plan["aggregation"]
        measure = plan.get("measure") or ""
        groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            groups[_display_cell(row.get(group_by)) if group_by else "全部"].append(row)
        value_label = "计数" if aggregation == "count" else f"{measure} · {aggregation}"
        output: list[dict[str, Any]] = []
        for key, members in groups.items():
            numbers = [
                number for member in members
                if (number := _numeric(member.get(measure))) is not None
            ]
            if aggregation == "count":
                value: Any = len(members)
            elif not numbers:
                value = ""
            elif aggregation == "sum":
                value = sum(numbers)
            elif aggregation == "avg":
                value = fmean(numbers)
            elif aggregation == "min":
                value = min(numbers)
            else:
                value = max(numbers)
            row = {value_label: value}
            if group_by:
                row = {group_by: key, **row}
            output.append(row)
        output.sort(key=lambda row: _numeric(row.get(value_label)) or 0, reverse=True)
        columns = ([group_by] if group_by else []) + [value_label]
        return output, columns

    @staticmethod
    def _top(
        rows: Sequence[dict[str, Any]], plan: dict[str, Any]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        sort_by = plan["sort_by"]
        reverse = plan.get("order") != "asc"
        ordered = sorted(
            rows,
            key=lambda row: (
                _numeric(row.get(sort_by)) is not None,
                _numeric(row.get(sort_by)) or 0,
                _display_cell(row.get(sort_by)),
            ),
            reverse=reverse,
        )
        headers = [key for key in ordered[0] if not key.startswith("__")] if ordered else []
        return ordered, headers

    @staticmethod
    def _tabular(
        rows: Sequence[dict[str, Any]], columns: Sequence[str]
    ) -> tuple[list[dict[str, Any]], list[str]]:
        selected = list(columns)
        if not selected and rows:
            selected = [key for key in rows[0] if not key.startswith("__")]
        return list(rows), selected

    @staticmethod
    def _citation(
        manifest: dict[str, Any],
        sheet: dict[str, Any],
        rows: Sequence[dict[str, Any]],
    ) -> Citation | None:
        element_id = next(
            (str(row.get("__element_id") or "") for row in rows if row.get("__element_id")),
            "",
        )
        return Citation(
            label=f"{manifest.get('source_title', '')} · {sheet['name']}!{sheet['range']}",
            source_id=manifest["source_id"],
            element_id=element_id,
            location_label=f"{sheet['name']}!{sheet['range']}",
            quoted_span="电子表格确定性分析结果",
            source_file_name=manifest.get("source_file_name", ""),
        )


def spreadsheet_prompt_block(
    results: Sequence[SpreadsheetAnalysisResult], *, preview_rows: int, max_bytes: int
) -> tuple[str, dict[str, dict[str, Any]]]:
    """Render bounded spreadsheet receipts into the ordinary citable map."""
    lines: list[str] = []
    evidence: dict[str, dict[str, Any]] = {}

    def append_line(line: str) -> bool:
        separator = "\n" if lines else ""
        current_bytes = len("\n".join(lines).encode("utf-8"))
        if current_bytes + len((separator + line).encode("utf-8")) > max_bytes:
            return False
        lines.append(line)
        return True

    for index, result in enumerate(results, start=1):
        key = f"k{6000 + index}"
        headers = [column.name for column in result.columns]
        if not append_line(
            f"{key}: [spreadsheet] {result.source_title} · {result.sheet}!{result.range} "
            f"operation={result.operation}; scanned={result.coverage.scanned_rows}; "
            f"result_rows={result.coverage.total_rows}"
        ):
            break
        can_append_rows = not headers or append_line(" | ".join(headers))
        for row in result.rows[: max(1, preview_rows)]:
            if not can_append_rows or not append_line(
                " | ".join(row.cells.get(header, "") for header in headers)
            ):
                break
        citation = next((row.citation for row in result.rows if row.citation), None)
        element_id = citation.element_id if citation else ""
        evidence[key] = {
            "object_id": element_id or result.source_id,
            "object_type": "element" if element_id else "source",
            "name": f"{result.sheet}!{result.range}",
            "definition": "电子表格确定性分析结果",
            "snippet": lines[-1][:300],
            "source_id": result.source_id,
            "element_id": element_id,
            "source_title": result.source_title,
            "source_file_name": result.source_file_name,
            "location_label": f"{result.sheet}!{result.range}",
            "tier": "personal",
            "notebook_id": "",
            "relevance": 1.0,
            "knowhow": None,
        }
    return "\n".join(lines), evidence
