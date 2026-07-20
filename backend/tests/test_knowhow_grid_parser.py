import io

import pytest
from openpyxl import Workbook

from app.services.knowhow.grid_parser import (
    GridParseError,
    ParsedGrid,
    guess_kinds,
    parse_grid,
)


def _xlsx_bytes(sheets: list[list[list[str]]]) -> bytes:
    """Build real xlsx bytes with openpyxl. First list is the first sheet."""
    wb = Workbook()
    ws = wb.active
    for row in sheets[0]:
        ws.append(row)
    for extra_rows in sheets[1:]:
        extra = wb.create_sheet()
        for row in extra_rows:
            extra.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _assert_chinese_message(err: Exception) -> None:
    message = str(err)
    assert message
    assert any("一" <= ch <= "鿿" for ch in message)


# --- parse_grid: xlsx ---------------------------------------------------


def test_parse_grid_xlsx_basic():
    header = ["违例概念", "现象识别方法", "根因分析动作", "修复方法", "依赖工具"]
    data_row = ["天线效应", "金属密度超标告警", "多晶硅未及时接地释放电荷", "插入跳线二极管", "Calibre DRC"]
    data = _xlsx_bytes([[header, data_row]])

    grid = parse_grid("rules.xlsx", data)

    assert isinstance(grid, ParsedGrid)
    assert grid.columns == header
    assert grid.rows == [data_row]


def test_parse_grid_xlsm_suffix_uses_same_parser():
    header = ["a", "b"]
    data_row = ["1", "2"]
    data = _xlsx_bytes([[header, data_row]])

    grid = parse_grid("rules.xlsm", data)

    assert grid.columns == header
    assert grid.rows == [data_row]


def test_parse_grid_xlsx_reads_first_sheet_only():
    first_sheet = [["a", "b"], ["1", "2"]]
    second_sheet = [["x", "y"], ["9", "9"]]
    data = _xlsx_bytes([first_sheet, second_sheet])

    grid = parse_grid("rules.xlsx", data)

    assert grid.columns == ["a", "b"]
    assert grid.rows == [["1", "2"]]


def test_parse_grid_xlsx_none_cells_become_empty_string():
    wb = Workbook()
    ws = wb.active
    ws.append(["h1", "h2", "h3"])
    ws.append(["a1", None, "a3"])
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.rows == [["a1", "", "a3"]]


def test_parse_grid_xlsx_non_string_cells_are_stringified():
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "count", "ratio"])
    ws.append(["widget", 42, 3.5])
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.rows == [["widget", "42", "3.5"]]


# --- parse_grid: xlsx merged-cell expansion -------------------------------
# openpyxl's iter_rows(values_only=True) 只在合并区左上角返回值、其余单元
# 格返回 None。用户在 Excel 里做的合并（横向标题跨列、竖向分组头跨行）
# 不展开时会被解析成一大片空——表头会重复列名（多空串），数据会掉大量非
# 空行。grid_parser 的合并展开规则：把左上角值 fill 到该合并区覆盖的每一
# 个单元格，让解析器看到的形状与人肉读 Excel 一致。


def test_parse_grid_xlsx_horizontal_merge_in_header_expands():
    # 表头一横排合并（比如「上传时间」下压合并成一个大标题覆盖 3 列）：
    # 展开后每列都拿到相同的表头名——`_build_grid` 的「表头重复列名」守卫
    # 仍然会驳回它（这才是正确行为，防止用户误传合并的表头当成多列），
    # 本用例只验证「合并被 fill 了、没有静默空串列」。
    wb = Workbook()
    ws = wb.active
    ws.append(["date", None, None])
    ws.append(["2024-01", "2024-02", "2024-03"])
    ws.merge_cells("A1:C1")
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(GridParseError) as info:
        parse_grid("rules.xlsx", buf.getvalue())
    # 保证不是「空列名」错误——合并已展开成同名列，才走「重复列名」分支。
    assert "重复" in str(info.value)
    assert "date" in str(info.value)


def test_parse_grid_xlsx_horizontal_merge_in_data_row_fills_across():
    # 数据行里横向合并——例如一个案例的某个字段值太长，用户在 Excel 里
    # 把它跨列写；展开后该值应出现在覆盖的所有列上（下游 UI 消费时按列
    # 取值仍然拿得到，不因合并而丢失）。
    wb = Workbook()
    ws = wb.active
    ws.append(["violation", "phase-a", "phase-b", "phase-c"])
    ws.append(["hold&setup", "shared_root_cause", None, None])
    ws.merge_cells("B2:D2")
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.columns == ["violation", "phase-a", "phase-b", "phase-c"]
    assert grid.rows == [["hold&setup", "shared_root_cause", "shared_root_cause", "shared_root_cause"]]


def test_parse_grid_xlsx_vertical_merge_in_column_fills_down():
    # 竖向合并——典型来自用户表的「转置后」形态：多行案例共享同一个
    # 「违例概念」值（原表是横向合并做主题头，用户手动转置后变成竖向
    # 合并）。展开后 4 行都应拿到同样的概念名，UI 才能把它们当成同族案
    # 例正常显示。
    wb = Workbook()
    ws = wb.active
    ws.append(["violation", "symptom", "fix"])
    ws.append(["hold&setup", "case-A symptom", "case-A fix"])
    ws.append([None, "case-B symptom", "case-B fix"])
    ws.append([None, "case-C symptom", "case-C fix"])
    ws.append([None, "case-D symptom", "case-D fix"])
    ws.merge_cells("A2:A5")
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.columns == ["violation", "symptom", "fix"]
    assert grid.rows == [
        ["hold&setup", "case-A symptom", "case-A fix"],
        ["hold&setup", "case-B symptom", "case-B fix"],
        ["hold&setup", "case-C symptom", "case-C fix"],
        ["hold&setup", "case-D symptom", "case-D fix"],
    ]


def test_parse_grid_xlsx_multiple_merges_coexist():
    # 一张表里同时存在多个合并区（横+竖）——展开互不干扰，每个区各自
    # 按左上角 fill 自己的覆盖范围。
    wb = Workbook()
    ws = wb.active
    ws.append(["group", "a", "b", "c"])
    ws.append(["G1", "x1", None, None])  # B2:D2 横向合并
    ws.append([None, "x2", "y2", "z2"])  # A2:A3 竖向合并
    ws.merge_cells("B2:D2")
    ws.merge_cells("A2:A3")
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.columns == ["group", "a", "b", "c"]
    assert grid.rows == [
        ["G1", "x1", "x1", "x1"],
        ["G1", "x2", "y2", "z2"],
    ]


def test_parse_grid_xlsx_rectangular_merge_fills_all_cells():
    # 矩形合并区（跨多行且跨多列）——覆盖范围内每一格都填左上角值，
    # 不只是边缘或对角线。
    wb = Workbook()
    ws = wb.active
    ws.append(["h1", "h2", "h3", "h4"])
    ws.append(["a", "shared", None, None])
    ws.append(["b", None, None, "b4"])
    ws.append(["c", None, None, "c4"])
    ws.merge_cells("B2:C4")  # 3 行 × 2 列
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.rows == [
        ["a", "shared", "shared", ""],
        ["b", "shared", "shared", "b4"],
        ["c", "shared", "shared", "c4"],
    ]


# --- parse_grid: csv ------------------------------------------------------


def test_parse_grid_csv_basic_with_bom():
    text = "h1,h2\na,b\n"
    data = text.encode("utf-8-sig")  # BOM-prefixed, must decode cleanly

    grid = parse_grid("rules.csv", data)

    assert grid.columns == ["h1", "h2"]
    assert grid.rows == [["a", "b"]]


def test_parse_grid_csv_row_length_mismatch_pads_and_truncates():
    text = "h1,h2,h3\na1,a2\nb1,b2,b3,b4\n"
    data = text.encode("utf-8")

    grid = parse_grid("rules.csv", data)

    assert grid.columns == ["h1", "h2", "h3"]
    assert grid.rows == [["a1", "a2", ""], ["b1", "b2", "b3"]]


# --- parse_grid: markdown --------------------------------------------------


def test_parse_grid_md_basic():
    text = (
        "| 违例概念 | 现象识别方法 |\n"
        "| --- | --- |\n"
        "| 天线效应 | 金属密度超标 |\n"
    )

    grid = parse_grid("rules.md", text.encode("utf-8"))

    assert grid.columns == ["违例概念", "现象识别方法"]
    assert grid.rows == [["天线效应", "金属密度超标"]]


def test_parse_grid_markdown_suffix_alias():
    text = "| a | b |\n| --- | --- |\n| 1 | 2 |\n"

    grid = parse_grid("rules.markdown", text.encode("utf-8"))

    assert grid.columns == ["a", "b"]
    assert grid.rows == [["1", "2"]]


def test_parse_grid_md_drops_alignment_separator_variants():
    text = (
        "| A | B | C |\n"
        "| :--- | ---: | :---: |\n"
        "| 1 | 2 | 3 |\n"
    )

    grid = parse_grid("rules.md", text.encode("utf-8"))

    assert grid.columns == ["A", "B", "C"]
    assert grid.rows == [["1", "2", "3"]]


def test_parse_grid_md_escaped_pipe_and_br():
    text = (
        "| 概念 | 说明 |\n"
        "| --- | --- |\n"
        "| A\\|B | line1<br>line2<br/>line3 |\n"
    )

    grid = parse_grid("rules.md", text.encode("utf-8"))

    assert grid.columns == ["概念", "说明"]
    assert grid.rows == [["A|B", "line1\nline2\nline3"]]


def test_parse_grid_md_ignores_non_pipe_lines():
    text = (
        "# Title\n"
        "Some prose before the table.\n"
        "| h1 | h2 |\n"
        "| --- | --- |\n"
        "| a | b |\n"
        "Trailing prose.\n"
    )

    grid = parse_grid("rules.md", text.encode("utf-8"))

    assert grid.columns == ["h1", "h2"]
    assert grid.rows == [["a", "b"]]


def test_parse_grid_md_row_length_mismatch_pads_and_truncates():
    text = (
        "| h1 | h2 | h3 |\n"
        "| --- | --- | --- |\n"
        "| a1 | a2 |\n"
        "| b1 | b2 | b3 | b4 |\n"
    )

    grid = parse_grid("rules.md", text.encode("utf-8"))

    assert grid.columns == ["h1", "h2", "h3"]
    assert grid.rows == [["a1", "a2", ""], ["b1", "b2", "b3"]]


@pytest.mark.parametrize(
    ("filename", "data"),
    [
        (
            "rules.xlsx",
            _xlsx_bytes(
                [[
                    ["属性", "记录 A", "记录 B"],
                    ["现象", "A 现象", "B 现象"],
                    ["根因", "A 根因", "B 根因"],
                ]]
            ),
        ),
        (
            "rules.xlsm",
            _xlsx_bytes(
                [[
                    ["属性", "记录 A", "记录 B"],
                    ["现象", "A 现象", "B 现象"],
                    ["根因", "A 根因", "B 根因"],
                ]]
            ),
        ),
        (
            "rules.csv",
            "属性,记录 A,记录 B\n现象,A 现象,B 现象\n根因,A 根因,B 根因\n".encode("utf-8"),
        ),
        (
            "rules.md",
            (
                "| 属性 | 记录 A | 记录 B |\n"
                "| --- | --- | --- |\n"
                "| 现象 | A 现象 | B 现象 |\n"
                "| 根因 | A 根因 | B 根因 |\n"
            ).encode("utf-8"),
        ),
    ],
)
def test_parse_grid_rows_orientation_transposes_to_column_attributes(filename, data):
    grid = parse_grid(filename, data, orientation="rows")

    assert grid.columns == ["属性", "现象", "根因"]
    assert grid.rows == [
        ["记录 A", "A 现象", "A 根因"],
        ["记录 B", "B 现象", "B 根因"],
    ]


def test_parse_grid_rows_orientation_pads_ragged_rows_before_transpose():
    data = (
        "属性,记录 A,记录 B\n"
        "现象,A 现象\n"
        "根因,A 根因,B 根因\n"
    ).encode("utf-8")

    grid = parse_grid("rules.csv", data, orientation="rows")

    assert grid.columns == ["属性", "现象", "根因"]
    assert grid.rows == [
        ["记录 A", "A 现象", "A 根因"],
        ["记录 B", "", "B 根因"],
    ]


# --- validation errors ------------------------------------------------------


def test_parse_grid_empty_csv_raises_grid_parse_error():
    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", b"")
    assert isinstance(excinfo.value, ValueError)
    _assert_chinese_message(excinfo.value)


def test_parse_grid_md_without_any_table_raises_grid_parse_error():
    text = "just some prose, no pipe table here.\n"
    with pytest.raises(GridParseError):
        parse_grid("rules.md", text.encode("utf-8"))


def test_parse_grid_duplicate_column_names_raise():
    text = "h1,h1,h3\na,b,c\n"
    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", text.encode("utf-8"))
    _assert_chinese_message(excinfo.value)


def test_parse_grid_zero_data_rows_raise():
    text = "h1,h2,h3\n"
    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", text.encode("utf-8"))
    _assert_chinese_message(excinfo.value)


def test_parse_grid_unsupported_suffix_raises():
    with pytest.raises(GridParseError):
        parse_grid("rules.txt", b"whatever")


def test_parse_grid_rejects_unknown_orientation():
    with pytest.raises(GridParseError) as excinfo:
        parse_grid(
            "rules.csv",
            b"name,value\nA,1\n",
            orientation="diagonal",
        )
    assert str(excinfo.value) == "非法的属性排列方式"


@pytest.mark.parametrize(
    ("text", "message_fragment"),
    [
        ("属性,记录 A\n,A 现象\n", "空列名"),
        ("属性,记录 A\n现象,A\n现象,B\n", "重复列名"),
        ("属性\n现象\n根因\n", "没有数据行"),
    ],
)
def test_parse_grid_rows_orientation_reuses_normalized_header_validation(
    text, message_fragment
):
    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", text.encode("utf-8"), orientation="rows")
    assert message_fragment in str(excinfo.value)


def test_parse_grid_column_orientation_points_transposed_shape_to_ui_choice():
    data = "属性,,\n现象,A,B\n根因,C,D\n".encode("utf-8")

    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", data, orientation="columns")

    assert str(excinfo.value) == (
        "这张表看起来是属性按行排列，"
        "请返回并选择“属性按行”后重新导入。"
    )


def test_parse_grid_default_orientation_remains_columns():
    data = b"name,value\nA,1\n"
    assert parse_grid("rules.csv", data) == parse_grid(
        "rules.csv", data, orientation="columns"
    )


def test_parse_grid_csv_gbk_encoded_chinese_falls_back_and_decodes():
    # Chinese Excel exports CSV as GBK/ANSI by default, not UTF-8.
    text = "概念,说明\n违例,超标\n"
    data = text.encode("gbk")

    grid = parse_grid("rules.csv", data)

    assert grid.columns == ["概念", "说明"]
    assert grid.rows == [["违例", "超标"]]


def test_parse_grid_csv_undecodable_bytes_raise_friendly_error():
    # Fails both utf-8-sig and gbk decoding.
    data = b"\xff\xff\xff\xff"

    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", data)

    assert str(excinfo.value) == "无法识别文件编码，请将文件另存为 UTF-8 编码后重试"


def test_parse_grid_xlsx_corrupted_bytes_raise_friendly_error():
    data = b"this is not a real xlsx file, just garbage bytes"

    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.xlsx", data)

    assert str(excinfo.value) == "无法读取 Excel 文件，请确认文件未损坏且为 .xlsx 格式"


def test_parse_grid_xlsx_zero_worksheet_workbook_raises_friendly_error(monkeypatch):
    # openpyxl refuses to SAVE a worksheet-less workbook ("At least one sheet
    # must be visible"), so a real fixture can't be built — monkeypatch
    # load_workbook to hand back one whose .worksheets is empty. The guard must
    # turn the would-be bare IndexError into the friendly corrupt/invalid message.
    import openpyxl

    class _NoSheetWorkbook:
        worksheets: list = []

        def close(self):
            pass

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **k: _NoSheetWorkbook())

    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.xlsx", b"PK\x03\x04 pretend-valid-zip")

    assert str(excinfo.value) == "无法读取 Excel 文件，请确认文件未损坏且为 .xlsx 格式"


# --- all-empty-row dropping (across all three formats) ----------------------


def test_parse_grid_xlsx_drops_trailing_all_empty_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["h1", "h2", "h3"])
    ws.append(["a1", "a2", "a3"])
    ws.append(["", "", ""])        # formatted-but-empty trailing row
    ws.append([None, None, None])  # truly blank trailing row
    buf = io.BytesIO()
    wb.save(buf)

    grid = parse_grid("rules.xlsx", buf.getvalue())

    assert grid.columns == ["h1", "h2", "h3"]
    # Both phantom trailing rows are dropped; total_rows stays honest.
    assert grid.rows == [["a1", "a2", "a3"]]


def test_parse_grid_csv_drops_all_empty_rows():
    text = "h1,h2\na,b\n,\n   ,   \n"
    grid = parse_grid("rules.csv", text.encode("utf-8"))

    assert grid.columns == ["h1", "h2"]
    assert grid.rows == [["a", "b"]]


def test_parse_grid_md_drops_all_empty_rows():
    text = (
        "| h1 | h2 |\n"
        "| --- | --- |\n"
        "| a | b |\n"
        "|   |   |\n"
    )
    grid = parse_grid("rules.md", text.encode("utf-8"))

    assert grid.columns == ["h1", "h2"]
    assert grid.rows == [["a", "b"]]


def test_parse_grid_all_blank_data_rows_raise_no_data_error():
    # Header present but every data row is blank -> nothing survives the drop.
    text = "h1,h2\n,\n   ,   \n"
    with pytest.raises(GridParseError) as excinfo:
        parse_grid("rules.csv", text.encode("utf-8"))
    _assert_chinese_message(excinfo.value)


# --- guess_kinds (PR-2+3 Task 1: behavior kinds + anchor suggestion) ---------


def test_guess_kinds_timing_fixup_five_columns():
    """The canonical time-series-fixup header: 识别/分析/修复 columns are all
    procedure now (one behavior kind, no sub-roles), 工具 is entity, and the
    违例概念 column is kind attribute but nominated as the anchor suggestion
    (概念 is an anchor-NAME keyword, not a content-kind keyword)."""
    columns = ["违例概念", "现象识别方法", "根因分析动作", "修复方法", "依赖工具"]
    kinds, anchor_idx = guess_kinds(columns)
    assert kinds == ["attribute", "procedure", "procedure", "procedure", "entity"]
    assert anchor_idx == 0


def test_guess_kinds_travel_log_gets_all_attribute_and_no_anchor():
    """The record-shaped table that motivated dropping the first-column
    fallback (design doc §① 旅行日志实证): nothing matches a kind keyword and
    no column NAME suggests an identity — every column stays attribute and
    the anchor suggestion is None, NOT column 0."""
    columns = ["日期", "出发地", "目的地", "交通", "住宿", "花费", "天气", "同行人", "备注"]
    kinds, anchor_idx = guess_kinds(columns)
    assert kinds == ["attribute"] * 9
    assert anchor_idx is None


def test_guess_kinds_anchor_keywords_pick_first_match_only():
    columns = ["工艺名称", "器件类型", "说明"]
    kinds, anchor_idx = guess_kinds(columns)
    assert kinds == ["attribute", "attribute", "attribute"]
    assert anchor_idx == 0  # 名称 hits first; the later 类型 hit is not consumed


def test_guess_kinds_english_anchor_keywords_and_casefold():
    columns = ["Violation Name", "Fix Steps", "Tools"]
    kinds, anchor_idx = guess_kinds(columns)
    # "Fix Steps"/"Tools" carry no CHINESE kind keyword (the kind vocabulary
    # is deliberately the brief's exact list); the anchor list does include
    # English identity words, matched casefolded.
    assert anchor_idx == 0
    assert kinds[0] == "attribute"


def test_guess_kinds_procedure_beats_entity_on_double_hit():
    # 修复(procedure) and 工具(entity) both present: procedure wins (listed
    # first — same priority-by-order convention the old guess_roles used).
    kinds, anchor_idx = guess_kinds(["修复工具"])
    assert kinds == ["procedure"]
    assert anchor_idx is None


def test_guess_kinds_kind_and_anchor_are_independent():
    """A column can match BOTH a kind keyword and an anchor-name keyword —
    the kind guess and the anchor suggestion are separate outputs (anchor is
    not a kind), so it keeps its content kind AND gets nominated."""
    kinds, anchor_idx = guess_kinds(["分析方法名称", "备注"])
    assert kinds == ["procedure", "attribute"]
    assert anchor_idx == 0


# --- forward_fill_column（anchor 分组列填充）-------------------------------
from app.services.knowhow.grid_parser import forward_fill_column


def test_forward_fill_column_fills_blanks_below_first_value():
    # 分组列"只写一次"：首行有值，后续空 → 全部继承。
    rows = [["hold&setup", "case-A"], ["", "case-B"], ["", "case-C"]]
    assert forward_fill_column(rows, 0) == [
        ["hold&setup", "case-A"],
        ["hold&setup", "case-B"],
        ["hold&setup", "case-C"],
    ]


def test_forward_fill_column_restarts_at_next_nonempty():
    # 多概念分段：每遇到新非空值就换继承源。
    rows = [["A", "1"], ["", "2"], ["B", "3"], ["", "4"]]
    assert forward_fill_column(rows, 0) == [
        ["A", "1"], ["A", "2"], ["B", "3"], ["B", "4"],
    ]


def test_forward_fill_column_leading_blanks_stay_blank():
    # 开头就空（前面无非空可继承）→ 保持空。
    rows = [["", "1"], ["A", "2"], ["", "3"]]
    assert forward_fill_column(rows, 0) == [["", "1"], ["A", "2"], ["A", "3"]]


def test_forward_fill_column_only_target_column():
    # 只填目标列；其他列的空是真空，不动。
    rows = [["A", "x"], ["", ""], ["", "z"]]
    assert forward_fill_column(rows, 0) == [["A", "x"], ["A", ""], ["A", "z"]]


def test_forward_fill_column_does_not_mutate_input():
    rows = [["A", "1"], ["", "2"]]
    forward_fill_column(rows, 0)
    assert rows == [["A", "1"], ["", "2"]]  # 原 list 不变


def test_forward_fill_column_whitespace_only_counts_as_blank():
    # 纯空白视为空（与 _build_grid 的 strip 判空一致）。
    rows = [["A", "1"], ["   ", "2"]]
    assert forward_fill_column(rows, 0) == [["A", "1"], ["A", "2"]]


# --- 属性行表探测：用户若仍以默认“属性按列”解析属性行表，错误提示应引导
# 回到产品内切换方向，不再要求去 Excel 手工转置。


def test_parse_grid_transposed_table_hints_to_choose_rows_orientation():
    # 用户的 EDA 违例表原貌：字段名在第一列、每列一个分支，表头行除首格
    # 外只有第一个分支有值 → 解析成多个空列名。要给产品内方向指引而非
    # 「重复列名 ''」或 Excel 手工操作。
    data = _xlsx_bytes(
        [
            [
                ["违例概念", "hold和setup打架", None, None],
                ["现象识别方法", "单条path跨多corner", None, None],
                ["根因分析动作", "path中个别inst", "cell delay占比", "noise同时吃"],
                ["修复方法", "换VT开窗", "加shielding", "提高vih"],
            ]
        ]
    )
    with pytest.raises(GridParseError) as info:
        parse_grid("know-how沉淀.xlsx", data)
    msg = str(info.value)
    assert "属性按行" in msg
    assert "Excel" not in msg


def test_parse_grid_blank_header_without_transposed_shape_asks_to_fill_names():
    # 表头有空列名但不是转置形状（首列自己就有空 → 不是字段名列）：
    # 提示补齐列名，别误导用户去转置。
    data = _xlsx_bytes(
        [
            [
                ["名称", None, None],
                ["a", "1", "2"],
                [None, "3", "4"],
            ]
        ]
    )
    with pytest.raises(GridParseError) as info:
        parse_grid("t.xlsx", data)
    msg = str(info.value)
    assert "空列名" in msg
    assert "转置" not in msg
