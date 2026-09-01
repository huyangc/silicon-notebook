from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from app.core.config import Settings
from app.repositories.analysis_artifacts import AnalysisArtifactStore
from app.services.cancellation import AskCancelled
from app.services.spreadsheet_analysis import (
    SpreadsheetAnalysisService,
    spreadsheet_prompt_block,
)


class _EventLog:
    def __init__(self) -> None:
        self.events = []
        self.logger = SimpleNamespace(
            exception=lambda *args, **kwargs: None,
            warning=lambda *args, **kwargs: None,
        )

    def emit(self, event) -> None:
        self.events.append(event)


class _OfflinePlanner:
    configured = False


class _AggregatePlanner:
    configured = True

    def chat_json(self, *args, **kwargs) -> str:
        return json.dumps({
            "source_id": "src-1",
            "sheet": "Sales",
            "operation": "aggregate",
            "aggregation": "sum",
            "measure": "Amount",
            "group_by": "Region",
            "filters": [],
        })


class _CancelledPlanner:
    configured = True

    def chat_json(self, *args, **kwargs) -> str:
        raise AskCancelled("cancelled")


class _CapturingProfilePlanner:
    configured = True

    def __init__(self) -> None:
        self.prompt = ""

    def chat_json(self, messages, *args, **kwargs) -> str:
        self.prompt = messages[0]["content"]
        return json.dumps({
            "source_id": "src-1",
            "sheet": "Sales",
            "operation": "profile",
        })


class _AllColumnsFilterPlanner:
    configured = True

    def chat_json(self, *args, **kwargs) -> str:
        return json.dumps({
            "source_id": "src-1",
            "sheet": "Sales",
            "operation": "filter",
            "filters": [],
            "columns": [],
        })


class _TopPlanner:
    configured = True

    def chat_json(self, *args, **kwargs) -> str:
        return json.dumps({
            "source_id": "src-1",
            "sheet": "Sales",
            "operation": "top",
            "sort_by": "Amount",
            "order": "desc",
        })


class _EmptyCountPlanner:
    configured = True

    def chat_json(self, *args, **kwargs) -> str:
        return json.dumps({
            "source_id": "src-1",
            "sheet": "Sales",
            "operation": "aggregate",
            "aggregation": "count",
            "filters": [{"column": "Region", "operator": "eq", "value": "Missing"}],
        })


class _PartiallyInvalidFilterPlanner:
    configured = True

    def chat_json(self, *args, **kwargs) -> str:
        return json.dumps({
            "source_id": "src-1",
            "sheet": "Sales",
            "operation": "filter",
            "filters": [
                {"column": "Region", "operator": "eq", "value": "East"},
                {"column": "Missing", "operator": "eq", "value": "x"},
            ],
            "columns": ["Region", "Amount"],
        })


def _settings(tmp_path: Path) -> Settings:
    return Settings(
        _env_file=None,
        storage_dir=str(tmp_path / "storage"),
        event_log_enabled=False,
        llm_log_enabled=False,
    )


def _source(path: Path):
    return SimpleNamespace(
        id="src-1",
        notebook_id="nb-1",
        title="Sales workbook",
        type="xlsx",
        file_name="sales.xlsx",
        file_path=str(path),
        file_hash="abc",
        source_url="",
    )


def _workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Amount", "Owner"])
    sheet.append(["East", 10, "A"])
    sheet.append(["West", 20, "B"])
    sheet.append(["East", 40, "C"])
    workbook.save(path)


def test_compile_and_offline_profile_are_bounded_and_citable(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    artifacts = AnalysisArtifactStore(
        Path(settings.storage_dir), retention_days=30
    )
    service = SpreadsheetAnalysisService(
        artifacts=artifacts,
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )

    assert service.compile_source(
        _source(path),
        notebook_name="Notebook",
        owner_id="user-1",
        row_element_ids={
            ("Sales", 1): "el-1",
            ("Sales", 2): "el-2",
            ("Sales", 3): "el-3",
            ("Sales", 4): "el-4",
        },
    ) is True

    results, trace = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="分析这个 Excel 的数据质量和缺失情况",
        planner_client=_OfflinePlanner(),
    )
    assert trace is not None
    assert trace.step_type == "spreadsheet"
    [result] = results
    assert result.kind == "spreadsheet"
    assert result.operation == "profile"
    assert result.coverage.scanned_rows == 3
    assert [column.name for column in result.columns][:3] == ["列", "类型", "非空"]
    assert result.rows[0].citation is not None
    assert result.rows[0].citation.element_id == "el-2"


def test_planner_selects_whitelisted_grouped_aggregate(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={("Sales", 2): "el-2"},
    )

    results, _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="按 Region 汇总 Amount 的总和",
        planner_client=_AggregatePlanner(),
    )
    [result] = results
    assert result.operation == "aggregate"
    assert result.rows[0].cells == {"Region": "East", "Amount · sum": "50"}
    assert result.rows[1].cells == {"Region": "West", "Amount · sum": "20"}


@pytest.mark.parametrize(
    "oversized_field,oversized_value",
    [
        ("filters", [{"column": "Region", "operator": "eq", "value": "East"}] * 9),
        ("columns", ["Region"] * 21),
    ],
)
def test_planner_plan_over_protocol_limits_falls_back_without_partial_execution(
    tmp_path, oversized_field, oversized_value,
):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    class _OversizedPlanner:
        configured = True

        def chat_json(self, *args, **kwargs) -> str:
            plan = {
                "source_id": "src-1",
                "sheet": "Sales",
                "operation": "filter",
                "filters": [],
                "columns": ["Region"],
            }
            plan[oversized_field] = oversized_value
            return json.dumps(plan)

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="筛选这个 Excel 的明细",
        planner_client=_OversizedPlanner(),
    )

    assert result.operation == "profile"


def test_top_sort_keeps_nonnumeric_values_last_in_both_directions():
    rows = [
        {"Amount": ""},
        {"Amount": "n/a"},
        {"Amount": 10},
        {"Amount": 2},
    ]

    ascending, _ = SpreadsheetAnalysisService._top(
        rows, {"sort_by": "Amount", "order": "asc"}
    )
    descending, _ = SpreadsheetAnalysisService._top(
        rows, {"sort_by": "Amount", "order": "desc"}
    )

    assert [row["Amount"] for row in ascending] == [2, 10, "", "n/a"]
    assert [row["Amount"] for row in descending] == [10, 2, "n/a", ""]

    text_descending, _ = SpreadsheetAnalysisService._top(
        [{"Owner": "A"}, {"Owner": "C"}, {"Owner": "B"}],
        {"sort_by": "Owner", "order": "desc"},
    )
    assert [row["Owner"] for row in text_descending] == ["C", "B", "A"]


def test_numeric_header_labels_are_not_replaced_by_a_later_text_row(tmp_path):
    path = tmp_path / "year-columns.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", 2024, 2025])
    sheet.append(["North", "High", "Low"])
    sheet.append(["South", "Low", "High"])
    workbook.save(path)
    settings = _settings(tmp_path)
    artifacts = AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30)
    service = SpreadsheetAnalysisService(
        artifacts=artifacts,
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )

    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )
    manifest = artifacts.load_spreadsheet_manifest("nb-1", "src-1")

    assert manifest is not None
    assert manifest["sheets"][0]["headers"] == ["Region", "2024", "2025"]
    assert manifest["sheets"][0]["rows"][0]["cells"] == ["North", "High", "Low"]


def test_empty_ungrouped_count_returns_zero(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="筛选不存在的 Region 后统计这个 Excel 的行数",
        planner_client=_EmptyCountPlanner(),
    )

    assert result.operation == "aggregate"
    assert result.rows[0].cells == {"计数": "0"}


def test_top_result_citation_anchors_the_first_sorted_row(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={
            ("Sales", 2): "el-east-10",
            ("Sales", 3): "el-west-20",
            ("Sales", 4): "el-east-40",
        },
    )

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="列出这个 Excel 中 Amount 最大的行",
        planner_client=_TopPlanner(),
    )

    assert result.rows[0].cells["Amount"] == "40"
    assert result.rows[0].citation is not None
    assert result.rows[0].citation.element_id == "el-east-40"


def test_partially_invalid_planner_filter_falls_back_instead_of_dropping_it(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="筛选这个 Excel 的 East 明细",
        planner_client=_PartiallyInvalidFilterPlanner(),
    )

    assert result.operation == "profile"
    assert result.coverage.scanned_rows == 3


def test_empty_filter_result_preserves_requested_table_columns():
    rows, columns = SpreadsheetAnalysisService._tabular(
        [], [], ["Region", "Amount", "Owner"]
    )

    assert rows == []
    assert columns == ["Region", "Amount", "Owner"]


def test_aggregate_keeps_empty_numeric_groups_after_negative_values():
    rows, columns = SpreadsheetAnalysisService._aggregate(
        [
            {"Region": "Measured", "Amount": -10},
            {"Region": "Missing", "Amount": "n/a"},
        ],
        {"group_by": "Region", "aggregation": "sum", "measure": "Amount"},
    )

    assert columns == ["Region", "Amount · sum"]
    assert rows == [
        {"Region": "Measured", "Amount · sum": -10.0},
        {"Region": "Missing", "Amount · sum": ""},
    ]


def test_analysis_loads_manifest_from_owning_participant_notebook(tmp_path):
    path = tmp_path / "base-sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    source = _source(path)
    source.id = "base-sheet"
    source.notebook_id = "base-kept"
    assert service.compile_source(
        source, notebook_name="Base", owner_id="owner-base", row_element_ids={}
    )

    [result], trace = service.analyze(
        notebook_id="nb-1",
        source_ids=(),
        source_refs=(("base-kept", "base-sheet"),),
        question="分析挂载参考库中的 Excel 数据概况",
        planner_client=_OfflinePlanner(),
    )

    assert result.source_id == "base-sheet"
    assert trace is not None


def test_result_preview_is_bounded_by_cells_and_serialized_bytes(tmp_path):
    path = tmp_path / "wide.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append([f"C{column}" for column in range(8)])
    for index in range(3):
        sheet.append([f"r{index}c{column}-{'v' * 300}" for column in range(8)])
    workbook.save(path)
    settings = Settings(
        _env_file=None,
        storage_dir=str(tmp_path / "storage"),
        event_log_enabled=False,
        llm_log_enabled=False,
        spreadsheet_analysis_result_cells=10,
        spreadsheet_analysis_result_bytes=1_024,
    )
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="筛选这个 Excel 的全部明细",
        planner_client=_AllColumnsFilterPlanner(),
    )
    table_payload = {
        "columns": [column.model_dump(mode="json") for column in result.columns],
        "rows": [row.model_dump(mode="json") for row in result.rows],
    }
    assert len(result.columns) * (len(result.rows) + 1) <= 10
    assert len(json.dumps(
        table_payload, ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")) <= 1_024
    assert result.coverage.complete is False
    assert result.operation == "filter"
    assert result.coverage.truncated_reason == "payload_limit"
    assert any("仅返回前" in warning for warning in result.warnings)

    prompt, evidence = spreadsheet_prompt_block(
        [result], preview_rows=100, max_bytes=300
    )
    assert len(prompt.encode("utf-8")) <= 300
    assert evidence
    _, full_evidence = spreadsheet_prompt_block(
        [result], preview_rows=100, max_bytes=65_536
    )
    assert "v" * 300 in full_evidence["k6001"]["snippet"]
    assert len(full_evidence["k6001"]["snippet"]) > 300


def test_spreadsheet_output_and_prompt_budgets_are_validated(tmp_path):
    invalid_values = {
        "spreadsheet_analysis_result_cells": 9,
        "spreadsheet_analysis_result_bytes": 1_023,
        "spreadsheet_analysis_prompt_bytes": 1_023,
        "spreadsheet_analysis_planner_catalog_bytes": 1_023,
    }
    for field, value in invalid_values.items():
        with pytest.raises(ValidationError):
            Settings(
                _env_file=None,
                storage_dir=str(tmp_path / field),
                **{field: value},
            )


def test_planner_catalog_is_a_disclosed_bounded_projection(tmp_path):
    path = tmp_path / "catalog.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append([f"H{index}-{'x' * 80}" for index in range(40)])
    sheet.append(list(range(40)))
    workbook.save(path)
    settings = Settings(
        _env_file=None,
        storage_dir=str(tmp_path / "storage"),
        event_log_enabled=False,
        llm_log_enabled=False,
        spreadsheet_analysis_planner_catalog_bytes=1_024,
    )
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )
    planner = _CapturingProfilePlanner()

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="分析这个 Excel 的数据概况",
        planner_client=planner,
    )
    catalog_json = planner.prompt.split("可用工作簿：", 1)[1]
    assert len(catalog_json.encode("utf-8")) <= 1_024
    assert any("规划目录已按内容量上限" in warning for warning in result.warnings)


def test_workbook_without_row_anchors_uses_source_level_citation(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    [result], _ = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="分析这个 Excel 的数据概况",
        planner_client=_OfflinePlanner(),
    )
    assert result.rows[0].citation is not None
    assert result.rows[0].citation.source_id == "src-1"
    assert result.rows[0].citation.element_id == ""
    _, evidence = spreadsheet_prompt_block(
        [result], preview_rows=20, max_bytes=65_536
    )
    assert evidence["k6001"]["object_id"] == "src-1"
    assert evidence["k6001"]["object_type"] == "source"


def test_ambiguous_single_characters_do_not_trigger_spreadsheet_planning(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    for question in (
        "银行政策是什么",
        "执行这个方案",
        "分析这个方案",
        "count the references in this PDF",
        "top competitors in this market",
        "group the arguments by theme",
    ):
        planner = _CapturingProfilePlanner()
        results, trace = service.analyze(
            notebook_id="nb-1",
            source_ids=("src-1",),
            question=question,
            planner_client=planner,
        )
        assert results == []
        assert trace is None
        assert planner.prompt == ""

    planner = _CapturingProfilePlanner()
    results, trace = service.analyze(
        notebook_id="nb-1",
        source_ids=("src-1",),
        question="sum Amount by Region",
        planner_client=planner,
    )
    assert results
    assert trace is not None
    assert planner.prompt


def test_filter_comparisons_preserve_zero_and_false_values():
    service = object.__new__(SpreadsheetAnalysisService)
    equals_zero = [{"column": "value", "operator": "eq", "value": 0}]
    equals_false = [{"column": "value", "operator": "eq", "value": False}]

    assert service._matches({"value": 0}, equals_zero)
    assert service._matches({"value": 0.0}, equals_zero)
    assert service._matches({"value": "0"}, equals_zero)
    assert not service._matches({"value": None}, equals_zero)
    assert not service._matches({"value": ""}, equals_zero)
    assert not service._matches({"value": False}, equals_zero)
    assert service._matches({"value": False}, equals_false)
    assert service._matches({"value": "FALSE"}, equals_false)
    assert not service._matches({"value": None}, equals_false)


def test_reasoning_spreadsheet_lane_honors_exclude_scope():
    from app.services.ask_service import AskService
    from app.services.source_scope import ActiveSourceScope

    class _CapturingAnalysis:
        def __init__(self) -> None:
            self.source_ids = ()
            self.source_refs = ()

        def analyze(self, **kwargs):
            self.source_ids = kwargs["source_ids"]
            self.source_refs = kwargs["source_refs"]
            return [], None

    analysis = _CapturingAnalysis()
    service = object.__new__(AskService)
    service.spreadsheet_analysis = analysis
    service.ask_engine_participant_notebooks = lambda notebook_id: (
        "nb-1", "base-kept", "base-excluded"
    )
    visible_by_notebook = {
        "nb-1": ("src-kept", "src-excluded", "src-hidden"),
        "base-kept": ("base-sheet",),
        "base-excluded": ("excluded-base-sheet",),
    }
    service.ask_engine_visible_sources = lambda notebook_id: visible_by_notebook[
        notebook_id
    ]
    service.ask_engine_hidden_sources = lambda notebook_id, user_id: ("src-hidden",)
    service.model_clients = SimpleNamespace(chat=lambda workload: _OfflinePlanner())
    runtime = SimpleNamespace(
        scope=ActiveSourceScope(
            notebook_id="nb-1",
            mode="exclude",
            source_ids=frozenset({"src-excluded"}),
            base_mode="include",
            base_notebook_ids=frozenset({"base-kept"}),
        ),
        cancellation=None,
        trace_sink=None,
    )
    prepared = SimpleNamespace(
        notebook_id="nb-1",
        user_id="user-1",
        research_question="按地区汇总销售额",
    )

    assert service._spreadsheet_reasoning_results(prepared, runtime, []) == []
    assert analysis.source_ids == ("src-kept",)
    assert analysis.source_refs == (
        ("nb-1", "src-kept"),
        ("base-kept", "base-sheet"),
    )


def test_planner_cancellation_is_not_downgraded_to_local_profile(tmp_path):
    path = tmp_path / "sales.xlsx"
    _workbook(path)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )
    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    )

    with pytest.raises(AskCancelled):
        service.analyze(
            notebook_id="nb-1",
            source_ids=("src-1",),
            question="分析这个 Excel",
            planner_client=_CancelledPlanner(),
        )


def test_challenging_workbook_is_archived_instead_of_silently_misanalysed(tmp_path):
    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Amount"])
    sheet.append(["East", 10])
    sheet.append([])
    sheet.append(["Owner", "Quota"])
    sheet.append(["A", 20])
    workbook.save(path)
    settings = _settings(tmp_path)
    artifacts = AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30)
    service = SpreadsheetAnalysisService(
        artifacts=artifacts,
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )

    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    ) is False
    [issue] = artifacts.list_issues(status="open")
    assert issue["code"] == "SPREADSHEET_MULTIPLE_REGIONS"
    assert issue["artifact_available"] is True
    assert artifacts.load_spreadsheet_manifest("nb-1", "src-1") is None


def test_oversized_cell_is_rejected_without_truncation(tmp_path):
    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Notes"])
    sheet.append(["East", "x" * 300])
    workbook.save(path)
    settings = Settings(
        _env_file=None,
        storage_dir=str(tmp_path / "storage"),
        event_log_enabled=False,
        llm_log_enabled=False,
        spreadsheet_analysis_max_cell_chars=256,
    )
    artifacts = AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30)
    service = SpreadsheetAnalysisService(
        artifacts=artifacts,
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )

    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    ) is False
    [issue] = artifacts.list_issues(status="open")
    assert issue["code"] == "SPREADSHEET_CELL_TOO_LONG"


def test_missing_header_is_rejected_instead_of_synthesized(tmp_path):
    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", None, "Owner"])
    sheet.append(["East", 10, "A"])
    workbook.save(path)
    settings = _settings(tmp_path)
    artifacts = AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30)
    service = SpreadsheetAnalysisService(
        artifacts=artifacts,
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )

    assert service.compile_source(
        _source(path), notebook_name="Notebook", owner_id="user-1",
        row_element_ids={},
    ) is False
    [issue] = artifacts.list_issues(status="open")
    assert issue["code"] == "SPREADSHEET_HEADER_MISSING"


def test_legacy_xls_preserves_date_boolean_error_and_number_types(
    tmp_path, monkeypatch
):
    class _Cell:
        def __init__(self, ctype, value):
            self.ctype = ctype
            self.value = value

    class _Sheet:
        name = "Legacy"
        nrows = 2
        ncols = 4

        @staticmethod
        def row(index):
            if index == 0:
                return [
                    _Cell(1, value)
                    for value in ("Date", "Flag", "Error", "Number")
                ]
            return [
                _Cell(3, 46024.0),
                _Cell(4, 1),
                _Cell(5, 42),
                _Cell(2, 3.5),
            ]

        @staticmethod
        def row_values(index):
            raise AssertionError("typed XLS cells must not be reduced to raw values")

    class _Book:
        nsheets = 1
        datemode = 0

        def __init__(self):
            self.sheet = _Sheet()
            self.released = False

        def sheet_by_index(self, index):
            assert index == 0
            return self.sheet

        def unload_sheet(self, index):
            assert index == 0

        def release_resources(self):
            self.released = True

    book = _Book()
    fake_xlrd = SimpleNamespace(
        XL_CELL_EMPTY=0,
        XL_CELL_TEXT=1,
        XL_CELL_NUMBER=2,
        XL_CELL_DATE=3,
        XL_CELL_BOOLEAN=4,
        XL_CELL_ERROR=5,
        XL_CELL_BLANK=6,
        error_text_from_code={42: "#N/A"},
        open_workbook=lambda *args, **kwargs: book,
        xldate_as_datetime=lambda value, datemode: datetime(2026, 1, 2, 3, 4, 5),
    )
    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)
    settings = _settings(tmp_path)
    service = SpreadsheetAnalysisService(
        artifacts=AnalysisArtifactStore(Path(settings.storage_dir), retention_days=30),
        settings=settings,
        event_log=_EventLog(),
        now=lambda: "2026-08-31T01:00:00+00:00",
    )

    [manifest] = service._compile_xls(tmp_path / "legacy.xls", {})

    assert manifest["rows"][0]["cells"] == [
        "2026-01-02T03:04:05",
        True,
        "#N/A",
        3.5,
    ]
    assert book.released is True


def test_issue_archive_resolves_redacts_and_expires(tmp_path):
    store = AnalysisArtifactStore(tmp_path, retention_days=2)
    source_path = tmp_path / "source.pdf"
    source_path.write_bytes(b"private")
    occurred = "2026-08-29T00:00:00+00:00"
    store.record_issue(
        notebook_id="nb-1",
        notebook_name="Private notebook",
        owner_id="user-1",
        source_id="src-1",
        source_title="Private title",
        file_name="private.pdf",
        source_type="pdf",
        category="source_parse",
        code="SOURCE_PARSE_FAILED",
        summary="safe",
        occurred_at=occurred,
        source_path=str(source_path),
    )
    [open_issue] = store.list_issues(
        status="open", now=datetime(2026, 8, 30, tzinfo=timezone.utc)
    )
    assert open_issue["artifact_available"] is True

    store.resolve_issue(
        "nb-1", "src-1", "source_parse",
        resolved_at="2026-08-30T12:00:00+00:00",
    )
    [resolved] = store.list_issues(
        status="resolved", now=datetime(2026, 8, 30, 13, tzinfo=timezone.utc)
    )
    assert resolved["artifact_available"] is False

    store.redact_source(
        "nb-1", "src-1", occurred_at="2026-08-30T14:00:00+00:00"
    )
    [redacted] = store.list_issues(
        now=datetime(2026, 8, 30, 15, tzinfo=timezone.utc)
    )
    assert redacted["source_deleted"] is True
    assert redacted["owner_id"] == ""
    assert redacted["notebook_id"] == ""
    assert redacted["notebook_name"] == ""
    assert redacted["source_title"] == ""
    assert redacted["file_name"] == ""
    assert redacted["code"] == ""
    assert redacted["source_type"] == ""
    assert "nb-1" not in redacted["id"]
    assert "src-1" not in redacted["id"]
    assert not (store.root / "issues" / "nb-1").exists()
    [metadata_path] = list((store.root / "issues").glob("*/*/*/issue.json"))
    relative_path = str(metadata_path.relative_to(store.root / "issues"))
    assert "nb-1" not in relative_path
    assert "src-1" not in relative_path

    assert store.list_issues(
        now=datetime(2026, 9, 1, 1, tzinfo=timezone.utc)
    ) == []
    assert not (store.root / "issues" / "redacted").exists()


def test_expired_issue_removes_identifier_directories(tmp_path):
    store = AnalysisArtifactStore(tmp_path, retention_days=1)
    store.record_issue(
        notebook_id="nb-private",
        notebook_name="Private notebook",
        owner_id="user-private",
        source_id="src-private",
        source_title="Private source",
        file_name="private.xlsx",
        source_type="xlsx",
        category="spreadsheet_analysis",
        code="SPREADSHEET_INVALID_OOXML",
        summary="safe",
        occurred_at="2026-08-29T00:00:00+00:00",
        archive_file=False,
    )

    assert store.list_issues(
        now=datetime(2026, 8, 31, tzinfo=timezone.utc)
    ) == []
    assert not (store.root / "issues" / "nb-private").exists()


def test_notebook_delete_does_not_fail_after_artifact_redaction_error(
    tmp_path, monkeypatch, caplog
):
    from app.services import notebook_catalog

    class _Store:
        @staticmethod
        def delete_row_and_orphan_embeddings(notebook_id: str) -> list[str]:
            return []

    class _BrokenArtifacts:
        @staticmethod
        def redact_notebook(notebook_id: str, *, occurred_at: str) -> None:
            raise OSError("private filesystem detail")

    service = object.__new__(notebook_catalog.NotebookCatalogService)
    service._store = _Store()
    service._storage_dir = lambda: tmp_path
    service._analysis_artifacts = _BrokenArtifacts()
    service.get_notebook = lambda notebook_id: object()
    monkeypatch.setattr(
        notebook_catalog, "_delete_notebook_asset_dir", lambda *args: None
    )

    service.delete_notebook("nb-1")

    assert "analysis artifact redaction failed (OSError)" in caplog.text
    assert "private filesystem detail" not in caplog.text


def test_notebook_artifact_redaction_precedes_fallible_file_cleanup(
    tmp_path, monkeypatch
):
    from app.services import notebook_catalog

    redacted: list[tuple[str, str]] = []

    class _Store:
        @staticmethod
        def delete_row_and_orphan_embeddings(notebook_id: str) -> list[str]:
            return ["private-source"]

    class _Artifacts:
        @staticmethod
        def redact_notebook(notebook_id: str, *, occurred_at: str) -> None:
            redacted.append((notebook_id, occurred_at))

    service = object.__new__(notebook_catalog.NotebookCatalogService)
    service._store = _Store()
    service._storage_dir = lambda: tmp_path
    service._analysis_artifacts = _Artifacts()
    service.get_notebook = lambda notebook_id: object()

    def _fail_file_cleanup(file_path: str) -> None:
        raise OSError("private filesystem detail")

    monkeypatch.setattr(notebook_catalog, "_delete_source_file", _fail_file_cleanup)
    monkeypatch.setattr(
        notebook_catalog, "_delete_notebook_asset_dir", lambda *args: None
    )

    with pytest.raises(OSError, match="private filesystem detail"):
        service.delete_notebook("nb-private")

    assert len(redacted) == 1
    assert redacted[0][0] == "nb-private"
