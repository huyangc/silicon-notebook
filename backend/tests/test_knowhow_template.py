# backend/tests/test_knowhow_template.py
"""knowhow-tables PR-2+3 Task 6: Excel template download (GET .../template)
and append import (POST .../append, mode=preview|commit) — design doc §②
路B "Excel 模板往返".

Mirrors test_knowhow_editing_api.py's fixture conventions (TestClient +
register/login; the create-table wizard for setup; the existing /rows
editing endpoint to seed "already there" rows via plain HTTP — no direct
facade import needed anywhere in this file, since every fixture need here
already has an HTTP equivalent). Scheduler dispatch is verified by
monkeypatching `app.services.knowhow.api.get_scheduler` to a small counting
fake (knowhow_routes.py always does a fresh `knowhow_api.get_scheduler(repo)`
attribute lookup per call, so patching the module attribute intercepts it)
rather than polling for real background projection to settle — deterministic
and fast, and this task's brief calls for exactly that ("patch scheduler
counting").
"""
from __future__ import annotations

import io
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.services.knowhow import api as knowhow_api
from app.services.knowhow.grid_parser import parse_grid


def _xlsx_bytes(header: list[str], rows: list[list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _client(tmp_path, monkeypatch):
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 't.db'}")
    monkeypatch.setenv("SILICON_NOTEBOOK_STORAGE_DIR", str(tmp_path / "s"))
    monkeypatch.setenv("EVENT_LOG_ENABLED", "false")
    monkeypatch.setenv("LLM_LOG_ENABLED", "false")
    from app.main import app
    return TestClient(app)


def _login(client, username, password="pw123456"):
    client.post("/api/auth/register", json={"username": username, "password": password})
    tok = client.post("/api/auth/login", json={"username": username, "password": password}).json()["token"]
    return {"Authorization": f"Bearer {tok}"}


def _mk_notebook(client, headers, name="N"):
    return client.post("/api/notebooks", json={"name": name}, headers=headers).json()["id"]


DEFAULT_COLUMNS = (("现象", "attribute"), ("方法", "procedure"), ("工具", "entity"))


def _create_table(client, headers, nb, *, title="判别表", columns=DEFAULT_COLUMNS, anchor_index=0):
    body: dict = {"title": title, "columns": [{"name": n, "kind": k} for n, k in columns]}
    if anchor_index is not None:
        body["anchor_index"] = anchor_index
    resp = client.post(f"/api/notebooks/{nb}/knowhow", headers=headers, json=body)
    assert resp.status_code == 200, resp.text
    return resp.json()


def _add_row(client, headers, nb, table_id, cells):
    resp = client.post(
        f"/api/notebooks/{nb}/knowhow/{table_id}/rows", headers=headers, json={"cells": cells},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


def _append(client, headers, nb, table_id, header, rows, mode="preview", fname="append.xlsx"):
    data = _xlsx_bytes(header, rows)
    return client.post(
        f"/api/notebooks/{nb}/knowhow/{table_id}/append",
        headers=headers,
        files={"file": (fname, data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"mode": mode},
    )


class _FakeScheduler:
    def __init__(self):
        self.calls: list[str] = []

    def schedule(self, table_id):
        self.calls.append(table_id)


def _patch_scheduler(monkeypatch) -> _FakeScheduler:
    fake = _FakeScheduler()
    monkeypatch.setattr(knowhow_api, "get_scheduler", lambda repo: fake)
    return fake


def _decoded_filename_star(disposition: str) -> str:
    marker = "filename*=UTF-8''"
    idx = disposition.index(marker) + len(marker)
    return unquote(disposition[idx:])


# ===========================================================================
# GET /notebooks/{nb}/knowhow/{t}/template
# ===========================================================================


def test_template_round_trips_column_names_through_grid_parser(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002001")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)

    resp = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}/template", headers=owner_h)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    # Table title is Chinese (non-ASCII) — the plain filename= fallback must
    # stay ASCII-safe, but the real title is recoverable from filename*=.
    assert _decoded_filename_star(disposition) == f"{table['title']}-template.xlsx"

    grid = parse_grid("template.xlsx", resp.content)
    assert grid.columns == [c["name"] for c in table["columns"]]


def test_template_second_row_carries_kind_and_anchor_hints_and_freezes_panes(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002002")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)  # anchor_index=0 -> 现象 becomes anchor

    resp = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}/template", headers=owner_h)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    hints = [ws.cell(row=2, column=i + 1).value for i in range(3)]
    assert "行标题" in hints[0]  # 现象 is the anchor column
    assert "方法步骤" in hints[1]
    assert "工具" in hints[2]
    assert ws.freeze_panes == "A3"


def test_template_reflects_current_columns_not_a_stale_snapshot(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002003")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)

    add_resp = client.post(
        f"/api/notebooks/{nb}/knowhow/{table['id']}/columns", headers=owner_h,
        json={"name": "备注", "kind": "attribute"},
    )
    assert add_resp.status_code == 200, add_resp.text

    resp = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}/template", headers=owner_h)
    grid = parse_grid("template.xlsx", resp.content)
    assert grid.columns == ["现象", "方法", "工具", "备注"]


def test_template_unknown_table_404(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002004")
    nb = _mk_notebook(client, owner_h)
    resp = client.get(f"/api/notebooks/{nb}/knowhow/does-not-exist/template", headers=owner_h)
    assert resp.status_code == 404


def test_template_slash_in_title_is_percent_escaped_in_filename_star(tmp_path, monkeypatch):
    """A literal "/" in a table title must come out as %2F inside the RFC 5987
    filename*= value (quote's DEFAULT safe="/" would leave it raw — invalid
    attr-char, and inconsistent with the ASCII fallback branch, which strips
    "/") while still round-tripping back to the real title via unquote."""
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002005")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb, title="输入/输出对照表")

    resp = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}/template", headers=owner_h)
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    marker = "filename*=UTF-8''"
    encoded_name = disposition[disposition.index(marker) + len(marker):]
    assert "%2F" in encoded_name
    assert "/" not in encoded_name  # no raw slash anywhere in the encoded form
    assert _decoded_filename_star(disposition) == "输入/输出对照表-template.xlsx"


# ===========================================================================
# POST /notebooks/{nb}/knowhow/{t}/append (mode=preview)
# ===========================================================================


def test_append_preview_matches_by_name_all_columns_present(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002010")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲", "加阻尼", "示波器"], ["欠冲", "调走线", "万用表"]],
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["total_rows"] == 2
    assert body["rows_preview"] == [["过冲", "加阻尼", "示波器"], ["欠冲", "调走线", "万用表"]]
    assert body["unmatched_columns"] == []
    assert body["duplicate_titles"] == []


def test_append_preview_only_first_five_rows_shown_but_total_rows_is_exact(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002011")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)

    rows = [[f"现象{i}", "x", "y"] for i in range(7)]
    resp = _append(client, owner_h, nb, table["id"], ["现象", "方法", "工具"], rows)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["total_rows"] == 7
    assert len(body["rows_preview"]) == 5
    assert body["rows_preview"][0] == ["现象0", "x", "y"]


def test_append_preview_missing_column_becomes_blank_not_an_error(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002012")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)  # 现象/方法/工具

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法"],  # 工具 entirely missing from the file
        [["过冲", "加阻尼"]],
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["rows_preview"] == [["过冲", "加阻尼", ""]]
    assert body["unmatched_columns"] == []


def test_append_preview_extra_column_reported_and_dropped_from_preview(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002013")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具", "备注"],
        [["过冲", "加阻尼", "示波器", "多余信息"]],
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["unmatched_columns"] == ["备注"]
    assert body["rows_preview"] == [["过冲", "加阻尼", "示波器"]]


def test_append_preview_flags_duplicate_titles_against_existing_rows_only(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002014")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    _patch_scheduler(monkeypatch)
    col_ids = {c["name"]: c["id"] for c in table["columns"]}
    _add_row(client, owner_h, nb, table["id"], {col_ids["现象"]: "过冲问题"})

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲问题", "加阻尼", "示波器"], ["全新问题", "x", "y"]],
    )
    assert resp.status_code == 200, resp.text
    # Only row 0 collides with the pre-existing row's title; row 1 ("全新
    # 问题") is not flagged, and the two incoming rows never get compared
    # against EACH OTHER even though nothing here happens to collide there.
    assert resp.json()["duplicate_titles"] == [{"row_index": 0, "title": "过冲问题"}]


def test_append_preview_blank_titles_never_count_as_duplicates(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002015")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    _patch_scheduler(monkeypatch)
    col_ids = {c["name"]: c["id"] for c in table["columns"]}
    # Existing row with a BLANK anchor cell (never written).
    _add_row(client, owner_h, nb, table["id"], {col_ids["方法"]: "旧方法"})

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["", "新方法", "y"]],  # incoming row's anchor cell is also blank
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["duplicate_titles"] == []


def test_append_preview_anchorless_table_never_flags_duplicates(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002016")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb, anchor_index=None)
    _patch_scheduler(monkeypatch)
    col_ids = {c["name"]: c["id"] for c in table["columns"]}
    _add_row(client, owner_h, nb, table["id"], {col_ids["现象"]: "过冲问题"})

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲问题", "加阻尼", "示波器"]],
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["duplicate_titles"] == []


def test_append_preview_bad_file_400(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002017")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    resp = client.post(
        f"/api/notebooks/{nb}/knowhow/{table['id']}/append",
        headers=owner_h,
        files={"file": ("x.txt", b"not a grid", "text/plain")},
        data={"mode": "preview"},
    )
    assert resp.status_code == 400


def test_append_invalid_mode_400(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002018")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    resp = _append(
        client, owner_h, nb, table["id"], ["现象", "方法", "工具"], [["a", "b", "c"]],
        mode="bogus",
    )
    assert resp.status_code == 400


def test_append_unknown_table_404(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002019")
    nb = _mk_notebook(client, owner_h)
    resp = _append(client, owner_h, nb, "does-not-exist", ["现象"], [["a"]])
    assert resp.status_code == 404


# ===========================================================================
# POST /notebooks/{nb}/knowhow/{t}/append (mode=commit)
# ===========================================================================


def test_append_commit_inserts_rows_and_schedules_reprojection(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002030")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    fake_scheduler = _patch_scheduler(monkeypatch)

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲", "加阻尼", "示波器"], ["欠冲", "调走线", "万用表"]],
        mode="commit",
    )
    assert resp.status_code == 200, resp.text
    assert resp.json() == {"added": 2}
    assert fake_scheduler.calls == [table["id"]]

    detail = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}", headers=owner_h).json()
    assert len(detail["rows"]) == 2
    col_ids = {c["name"]: c["id"] for c in table["columns"]}
    values = {row["cells"].get(col_ids["现象"]) for row in detail["rows"]}
    assert values == {"过冲", "欠冲"}


def test_append_commit_records_one_import_append_with_import_origin_and_actor(
    tmp_path, monkeypatch,
):
    """knowhow 表版本管理 Task 13 code review: actor/origin threading through
    commit_append had ZERO HTTP-level test coverage (a mutation reverting the
    real ``actor=user.id``/``origin="import"`` to defaults left the entire
    4817-test suite green). Assert the real end-to-end shape: the batch of
    newly appended rows lands as ONE ``import_append`` flow entry, its
    ``actor`` is the logged-in user's own id (not empty), and its ``origin``
    is ``"import"`` (never the manual editor's ``"user"``)."""
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002035")
    owner_id = client.get("/api/me", headers=owner_h).json()["id"]
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲", "加阻尼", "示波器"], ["欠冲", "调走线", "万用表"]],
        mode="commit",
    )
    assert resp.status_code == 200, resp.text

    changes = client.get(
        f"/api/notebooks/{nb}/knowhow/{table['id']}/history", headers=owner_h,
    ).json()["changes"]
    import_change = next(c for c in changes if c["kind"] == "import_append")
    assert import_change["origin"] == "import"
    assert import_change["actor"] == owner_id
    assert import_change["actor"] != ""
    assert len(import_change["payload"]["rows"]) == 2


def test_append_commit_ignores_missing_and_extra_columns_same_as_preview(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002031")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    _patch_scheduler(monkeypatch)

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "备注"],  # missing 方法/工具 (table columns); extra 备注 (file column)
        [["过冲", "多余信息"]],
        mode="commit",
    )
    assert resp.status_code == 200, resp.text
    assert resp.json() == {"added": 1}

    detail = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}", headers=owner_h).json()
    col_ids = {c["name"]: c["id"] for c in table["columns"]}
    row = detail["rows"][0]
    assert row["cells"].get(col_ids["现象"]) == "过冲"
    # Missing-column cells are never written at all (no empty-string
    # placeholder row) — matches get_knowhow_table's own "no cell row = never
    # edited" contract.
    assert col_ids["方法"] not in row["cells"]
    assert col_ids["工具"] not in row["cells"]


def test_append_commit_still_succeeds_despite_duplicate_titles(tmp_path, monkeypatch):
    """preview's duplicate_titles is advisory only — commit must not silently
    block on it (design doc: "确认后追加导入" — the human already decided)."""
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002032")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    _patch_scheduler(monkeypatch)
    col_ids = {c["name"]: c["id"] for c in table["columns"]}
    _add_row(client, owner_h, nb, table["id"], {col_ids["现象"]: "过冲问题"})

    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲问题", "加阻尼", "示波器"]],
        mode="commit",
    )
    assert resp.status_code == 200, resp.text
    assert resp.json() == {"added": 1}

    detail = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}", headers=owner_h).json()
    assert len(detail["rows"]) == 2  # pre-existing row + the newly appended duplicate-titled one


def test_append_commit_normalizes_markdown_but_preserves_special_characters(tmp_path, monkeypatch):
    """knowhow-md-normalize Task 5: commit_append now runs every non-empty
    cell through rule_normalize (zero LLM) before storing, so this can no
    longer assert byte-for-byte verbatim storage (its pre-Task-5 name/claim).
    What must still hold — and is exactly what this test now pins — is that
    markdown special characters (`**`/`_`/backtick/`#`) survive untouched;
    the ONLY change rule_normalize makes to this already-clean input is
    inserting the blank line CommonMark wants between a prose paragraph and
    an immediately-following list (a formatting-only change: same characters,
    same line content, one extra blank separator — see
    test_md_normalize_rule.py/knowhow_normalize_golden.json for rule_normalize's
    own dedicated unit coverage of that transform)."""
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002033")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    _patch_scheduler(monkeypatch)
    col_ids = {c["name"]: c["id"] for c in table["columns"]}

    markdown_text = "**加粗** _斜体_ `代码` # 标题\n- 步骤一\n- 步骤二"
    resp = _append(
        client, owner_h, nb, table["id"],
        ["现象", "方法", "工具"],
        [["过冲", markdown_text, "示波器"]],
        mode="commit",
    )
    assert resp.status_code == 200, resp.text

    detail = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}", headers=owner_h).json()
    row = detail["rows"][0]
    # rule_normalize's only change here: a blank line inserted before the list.
    expected = "**加粗** _斜体_ `代码` # 标题\n\n- 步骤一\n- 步骤二"
    assert row["cells"][col_ids["方法"]] == expected


def test_append_commit_bad_file_400_and_writes_nothing(tmp_path, monkeypatch):
    client = _client(tmp_path, monkeypatch)
    owner_h = _login(client, "a00002034")
    nb = _mk_notebook(client, owner_h)
    table = _create_table(client, owner_h, nb)
    _patch_scheduler(monkeypatch)

    resp = client.post(
        f"/api/notebooks/{nb}/knowhow/{table['id']}/append",
        headers=owner_h,
        files={"file": ("x.txt", b"not a grid", "text/plain")},
        data={"mode": "commit"},
    )
    assert resp.status_code == 400

    detail = client.get(f"/api/notebooks/{nb}/knowhow/{table['id']}", headers=owner_h).json()
    assert detail["rows"] == []
